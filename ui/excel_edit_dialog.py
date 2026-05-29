import sys
import os
import json
import tempfile
from datetime import datetime
from copy import copy
from PySide6.QtCore import qInstallMessageHandler, QtMsgType, QTimer, QSettings, Qt, QEvent, QSize
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QFileDialog, QMessageBox, QMenu, QHeaderView,
    QComboBox, QWidget, QLabel, QAbstractItemView, QLineEdit, QInputDialog,
    QCheckBox, QToolButton, QStyledItemDelegate, QStyleOptionViewItem
)
from PySide6.QtWidgets import QApplication
from PySide6.QtGui import QAction, QBrush, QColor, QFont, QShortcut, QKeySequence
from PySide6.QtWidgets import QStyle

import qtawesome as qta

from utils.logger import get_logger

logger = get_logger()

def myMessageHandler(msg_type, context, message):
    msg_lower = message.lower()
    skip_patterns = [
        "unknown property",
        "box-shadow",
        "transition",
        "transform",
        "qfont::setpointsize"
    ]
    if any(p in msg_lower for p in skip_patterns):
        return
    sys.stderr.write(f"{message}\n")

qInstallMessageHandler(myMessageHandler)

ICON_SIZE_SMALL = QSize(16, 16)
ICON_SIZE_MEDIUM = QSize(20, 20)
ICON_SIZE_LARGE = QSize(24, 24)


class TableEditorDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setFrame(False)
        editor.setAlignment(Qt.AlignCenter)
        return editor

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class CustomHeaderView(QHeaderView):
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.setSectionsClickable(True)
        self.setSectionsMovable(True)

    def mouseDoubleClickEvent(self, event):
        pos = event.pos()
        logical_index = self.logicalIndexAt(pos)
        if logical_index != -1:
            self.resizeSection(logical_index, QHeaderView.ResizeToContents)
        super().mouseDoubleClickEvent(event)


class FindReplaceDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("查找与替换")
        self.setMinimumWidth(400)
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)
        
        self._apply_theme()

        find_layout = QHBoxLayout()
        find_layout.addWidget(QLabel("查找内容:"))
        self.find_edit = QLineEdit()
        self.find_edit.setStyleSheet(f"""
            QLineEdit {{
                border: 1px solid {self._border_color};
                border-radius: 6px;
                padding: 6px 10px;
                background-color: {self._bg_color};
                color: {self._text_color};
            }}
        """)
        find_layout.addWidget(self.find_edit)
        layout.addLayout(find_layout)

        replace_layout = QHBoxLayout()
        replace_layout.addWidget(QLabel("替换为:"))
        self.replace_edit = QLineEdit()
        self.replace_edit.setStyleSheet(f"""
            QLineEdit {{
                border: 1px solid {self._border_color};
                border-radius: 6px;
                padding: 6px 10px;
                background-color: {self._bg_color};
                color: {self._text_color};
            }}
        """)
        replace_layout.addWidget(self.replace_edit)
        layout.addLayout(replace_layout)

        options_layout = QHBoxLayout()
        self.case_sensitive = QCheckBox("区分大小写")
        self.whole_word = QCheckBox("全字匹配")
        checkbox_style = f"""
            QCheckBox {{
                color: {self._text_color};
                font-size: 13px;
            }}
        """
        self.case_sensitive.setStyleSheet(checkbox_style)
        self.whole_word.setStyleSheet(checkbox_style)
        options_layout.addWidget(self.case_sensitive)
        options_layout.addWidget(self.whole_word)
        layout.addLayout(options_layout)

        btn_layout = QHBoxLayout()
        self.find_next_btn = QPushButton("查找下一个")
        self.replace_btn = QPushButton("替换")
        self.replace_all_btn = QPushButton("全部替换")
        
        btn_style = f"""
            QPushButton {{
                padding: 6px 16px;
                background-color: {self._primary_color};
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {self._primary_hover};
            }}
            QPushButton:pressed {{
                background-color: {self._primary_color};
            }}
        """
        self.find_next_btn.setStyleSheet(btn_style)
        self.replace_btn.setStyleSheet(btn_style)
        self.replace_all_btn.setStyleSheet(btn_style)
        
        btn_layout.addWidget(self.find_next_btn)
        btn_layout.addWidget(self.replace_btn)
        btn_layout.addWidget(self.replace_all_btn)
        layout.addLayout(btn_layout)

        self.find_next_btn.clicked.connect(self.find_next)
        self.replace_btn.clicked.connect(self.replace_current)
        self.replace_all_btn.clicked.connect(self.replace_all)
        self.last_found = None
    
    def _apply_theme(self):
        theme = {
            'bg': '#ffffff',
            'card': '#ffffff',
            'text': '#303133',
            'border': '#dcdfe6',
            'primary': '#409eff',
            'primary_hover': '#66b1ff'
        }
        if self.parent and hasattr(self.parent, 'theme_manager'):
            try:
                theme = self.parent.theme_manager.get_theme()
            except:
                pass
        self._bg_color = theme.get('bg', '#ffffff')
        self._text_color = theme.get('text', '#303133')
        self._border_color = theme.get('border', '#dcdfe6')
        self._primary_color = theme.get('primary', '#409eff')
        self._primary_hover = theme.get('primary_hover', '#66b1ff')
        self.setStyleSheet(f"background-color: {theme.get('card', '#ffffff')};")

    def find_next(self):
        search_text = self.find_edit.text()
        if not search_text:
            QMessageBox.warning(self, "提示", "请输入查找内容")
            return False
        table = self.parent.table
        flags = Qt.MatchContains
        if self.case_sensitive.isChecked():
            flags |= Qt.MatchCaseSensitive
        if self.whole_word.isChecked():
            flags |= Qt.MatchExactly

        current = table.currentItem()
        start_row = current.row() if current else 0
        start_col = current.column() if current else 0

        items = table.findItems(search_text, flags)
        if items:
            for item in items:
                r, c = item.row(), item.column()
                if r > start_row or (r == start_row and c >= start_col):
                    table.setCurrentItem(item)
                    self.last_found = (r, c)
                    return True
            table.setCurrentItem(items[0])
            self.last_found = (items[0].row(), items[0].column())
            QMessageBox.information(self, "提示", "已循环到开头")
            return True
        else:
            QMessageBox.information(self, "提示", "找不到内容")
            return False

    def replace_current(self):
        if not self.find_next():
            return
        current = self.parent.table.currentItem()
        if current:
            old_text = current.text()
            new_text = self.replace_edit.text()
            if old_text == self.find_edit.text():
                current.setText(new_text)
                self.parent.push_undo_action(current.row(), current.column(), old_text, new_text)
                self.parent.modified = True
                self.parent.modified_label.setText("*")

    def replace_all(self):
        search_text = self.find_edit.text()
        replace_text = self.replace_edit.text()
        if not search_text:
            return
        table = self.parent.table
        flags = Qt.MatchContains
        if self.case_sensitive.isChecked():
            flags |= Qt.MatchCaseSensitive
        if self.whole_word.isChecked():
            flags |= Qt.MatchExactly
        items = table.findItems(search_text, flags)
        count = 0
        for item in items:
            if item.text() == search_text:
                item.setText(replace_text)
                self.parent.push_undo_action(item.row(), item.column(), search_text, replace_text)
                count += 1
        if count > 0:
            self.parent.modified = True
            self.parent.modified_label.setText("*")
            QMessageBox.information(self, "完成", f"共替换 {count} 处")
        else:
            QMessageBox.information(self, "提示", "没有找到匹配内容")


class ExcelEditDialog(QDialog):
    MAX_UNDO_STEPS = 100
    def __init__(self, parent=None, is_pro=True, theme_manager=None):
        super().__init__(parent)
        self.is_pro = is_pro
        self.theme_manager = theme_manager
        self.setWindowTitle("Excel数据编辑" + ("（PRO专属）" if is_pro else "（只读模式 - 免费版）"))
        self.setMinimumSize(1000, 600)
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint)

        self.file_path = None
        self.modified = False
        self.sheet_names = []
        self.current_sheet = 0
        self.language = "zh"
        self.is_dark_mode = False
        if self.theme_manager:
            dark_themes = ["dark", "pink", "dark_blue"]
            self.is_dark_mode = self.theme_manager.current_theme in dark_themes
        self.main_window = parent

        self.frozen_row = -1

        self.undo_stack = []
        self.redo_stack = []
        self._in_undo_redo = False
        self._editing_item = None
        self._editing_old_value = None

        self.now_sym = "double_quote"
        self.now_pos = "both"
        self.split_mode = 0
        if self.main_window and hasattr(self.main_window, 'split_mode'):
            self.split_mode = self.main_window.split_mode
        self.skip_empty_cells = self.main_window.skip_empty_cells if self.main_window else False

        self.cell_formula = dict()
        self.cell_result = dict()
        self._is_loading_data = False
        self.is_editing = False

        self.original_workbook = None
        self.merged_cells = []
        self.cell_styles = {}
        self.recent_files = []

        self.backup_timer = QTimer()
        self.backup_timer.timeout.connect(self.auto_backup)
        self.backup_timer.start(300000)
        self.backup_file = None

        self.zoom_factor = 1.0
        self.original_font_sizes = {}
        
        # 缩放基准值 - 在初始化/加载数据时设置，缩放不改变它们
        self._base_font_size = 10       # 基准字体大小（原始值）
        self._base_row_height = 25      # 基准行高
        self._base_col_width = 100      # 基准列宽

        self.edited_cells = set()

        self.init_ui()

        if not self.is_pro:
            self._set_readonly_mode()

    # ---------- 辅助方法（已清空，不再操作背景） ----------
    def _refresh_all_cell_styles(self):
        # 不再做任何背景/前景设置，完全依赖样式表
        pass

    # ---------- 其他方法 ----------
    def update_split_mode(self, split_mode):
        self.split_mode = split_mode
    
    def update_skip_empty_cells(self, skip_empty):
        self.skip_empty_cells = skip_empty

    def _build_symbol_submenu(self):
        cfg = self.get_symbol_config()
        if not cfg:
            return None
        sym_menu = QMenu("📝 符号处理", self)
        if self.theme_manager:
            theme = self.theme_manager.get_theme()
            sym_menu.setStyleSheet(f"QMenu {{ background-color: {theme['card']}; border: 1px solid {theme['border']}; border-radius: 8px; color: {theme['text']}; }}")

        sym_menu.addSeparator()

        default_syms = {k: v for k, v in cfg.items() if not k.startswith("custom_")}
        custom_syms = {k: v for k, v in cfg.items() if k.startswith("custom_")}

        for sym_id, (left, right, full, name) in default_syms.items():
            sub_menu = QMenu(f"{left}{right} {name}", sym_menu)
            for pos_name, pos_key in [("头部", "head"), ("尾部", "tail"), ("两端", "both"), ("分隔", "split")]:
                act = QAction(pos_name, sub_menu)
                act.triggered.connect(lambda checked, s=sym_id, p=pos_key: self.quick_process_with(s, p))
                sub_menu.addAction(act)
            sym_menu.addMenu(sub_menu)

        if custom_syms and self.is_pro:
            sym_menu.addSeparator()
            for sym_id, (left, right, full, name) in custom_syms.items():
                sub_menu = QMenu(f"⭐ {left}{right} {name}", sym_menu)
                for pos_name, pos_key in [("头部", "head"), ("尾部", "tail"), ("两端", "both"), ("分隔", "split")]:
                    act = QAction(pos_name, sub_menu)
                    act.triggered.connect(lambda checked, s=sym_id, p=pos_key: self.quick_process_with(s, p))
                    sub_menu.addAction(act)
                sym_menu.addMenu(sub_menu)
        return sym_menu

    def _set_readonly_mode(self):
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_readonly_context_menu)
        self.btn_add.setEnabled(False)
        self.btn_delete.setEnabled(False)
        self.btn_save.setEnabled(False)
        self.btn_freeze_row.setEnabled(False)
        self.btn_clean.setEnabled(False)
        self.btn_export_csv.setEnabled(False)
        self.btn_find.setEnabled(False)
        self.sheet_combo.setEnabled(False)
        self.formula_edit.setReadOnly(True)
        self.formula_edit.setPlaceholderText("编辑功能为 PRO 版专属")
        if hasattr(self, 'file_btn'):
            menu = self.file_btn.menu()
            if menu:
                for action in menu.actions():
                    if action.text() not in ["📂 打开..."]:
                        action.setEnabled(False)

    def _show_pro_warning(self):
        QMessageBox.warning(self, "免费版限制", "此为 PRO 版功能，请升级后使用。")

    def keyPressEvent(self, event):
        key = event.key()
        mod = event.modifiers()
        
        if not self.is_pro:
            if (mod & Qt.ControlModifier) and key == Qt.Key_C:
                self.copy_cell()
                return
            if key in (Qt.Key_Up, Qt.Key_Down, Qt.Key_Left, Qt.Key_Right,
                    Qt.Key_PageUp, Qt.Key_PageDown, Qt.Key_Home, Qt.Key_End):
                self.table.keyPressEvent(event)
                return
            return
        
        if key == Qt.Key_Z and (mod & Qt.ControlModifier):
            if mod & Qt.ShiftModifier:
                self.redo()
            else:
                self.undo()
            return
        
        if key in (Qt.Key_Enter, Qt.Key_Return):
            cur_r = self.table.currentRow()
            cur_c = self.table.currentColumn()
            row_cnt = self.table.rowCount()
            if cur_r < row_cnt - 1:
                self.table.setCurrentCell(cur_r + 1, cur_c)
            else:
                self.add_row()
                self.table.setCurrentCell(row_cnt, cur_c)
            return
        
        if key == Qt.Key_Tab:
            cur_r = self.table.currentRow()
            cur_c = self.table.currentColumn()
            col_cnt = self.table.columnCount()
            if mod & Qt.ShiftModifier:
                if cur_c > 0:
                    self.table.setCurrentCell(cur_r, cur_c - 1)
            else:
                if cur_c < col_cnt - 1:
                    self.table.setCurrentCell(cur_r, cur_c + 1)
                elif cur_r < self.table.rowCount() - 1:
                    self.table.setCurrentCell(cur_r + 1, 0)
            return
        
        if key in (Qt.Key_Backspace, Qt.Key_Delete):
            for item in self.table.selectedItems():
                old = item.data(Qt.UserRole) or item.text()
                r, c = self.table.row(item), self.table.column(item)
                self.push_undo_action(r, c, old, "")
                item.setText("")
                item.setData(Qt.UserRole, None)
                self.edited_cells.add((r, c))
                # 不再设置背景，样式表会自动处理
            self.modified = True
            self.modified_label.setText("*")
            return
        
        super().keyPressEvent(event)

    def _show_readonly_context_menu(self, pos):
        menu = QMenu()
        copy_action = menu.addAction("复制")
        copy_action.triggered.connect(self._copy_selected_cells)
        menu.addSeparator()
        upgrade_action = menu.addAction("⚠️ 升级 PRO 版解锁全部功能")
        upgrade_action.triggered.connect(self._show_upgrade_dialog)
        menu.exec_(self.table.mapToGlobal(pos))

    def _show_upgrade_dialog(self):
        if self.main_window and hasattr(self.main_window, '_show_activation_dialog'):
            self.main_window._show_activation_dialog()
        elif self.main_window and hasattr(self.main_window, 'show_toast'):
            self.main_window.show_toast("请联系客服获取 PRO 版激活码")

    def _copy_selected_cells(self):
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            return
        clipboard = QApplication.clipboard()
        text_lines = []
        for range_item in selected_ranges:
            for row in range(range_item.topRow(), range_item.bottomRow() + 1):
                row_cells = []
                for col in range(range_item.leftColumn(), range_item.rightColumn() + 1):
                    item = self.table.item(row, col)
                    row_cells.append(item.text() if item else "")
                text_lines.append("\t".join(row_cells))
        clipboard.setText("\n".join(text_lines))

    def _show_function_support_tip(self):
        tip_text = """
<h3>内置公式计算器支持（可直接得出计算结果）：</h3>
<p>✅ <b>基础数学</b>：SUM, AVERAGE, COUNT, MAX, MIN</p>
<p>✅ <b>条件函数</b>：IF（简化版：IF(条件, 真值, 假值)）</p>
<p>✅ <b>查找函数</b>：VLOOKUP（简化版：精确匹配）</p>
<p>⚠️ 其他函数（如 LEFT, RIGHT, TODAY, AND, OR 等）<b>不会自动计算</b>，将保留公式文本。</p>
<p>✅ 保存为 .xlsx 后用 Excel 打开，所有公式均可正常计算。</p>
"""
        QMessageBox.information(self, "📋 函数支持说明", tip_text)

    def _check_pro_edit_permission(self):
        if not self.is_pro:
            QMessageBox.warning(self, "权限不足", "编辑功能为 PRO 版专属，请升级到 PRO 版。")
            return False
        return True

    # ------------------- UI 初始化 -------------------
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        top_layout = QHBoxLayout()

        self.file_btn = QPushButton(" 文件")
        self.file_btn.setIcon(qta.icon('fa5s.folder', color='#606266'))
        self.file_btn.setIconSize(ICON_SIZE_MEDIUM)
        self.file_btn.setToolTip("文件操作（新建、打开、保存等）")
        self.file_btn.setAutoDefault(False)
        file_menu = QMenu(self)
        new_action = QAction(" 新建", self)
        new_action.setIcon(qta.icon('fa5s.file', color='#606266'))
        new_action.setToolTip("新建空白表格 (Ctrl+N)")
        new_action.triggered.connect(self.new_file)
        file_menu.addAction(new_action)
        open_action = QAction(" 打开...", self)
        open_action.setIcon(qta.icon('fa5s.folder-open', color='#606266'))
        open_action.setToolTip("打开 Excel 文件 (Ctrl+O)")
        open_action.triggered.connect(self.import_excel)
        file_menu.addAction(open_action)
        file_menu.addSeparator()
        save_action = QAction(" 保存", self)
        save_action.setIcon(qta.icon('fa5s.save', color='#606266'))
        save_action.setToolTip("保存当前文件 (Ctrl+S)")
        save_action.triggered.connect(self.save_excel)
        file_menu.addAction(save_action)
        save_as_action = QAction(" 另存为...", self)
        save_as_action.setIcon(qta.icon('fa5s.save', color='#606266'))
        save_as_action.setToolTip("另存为新的文件 (Ctrl+Shift+S)")
        save_as_action.triggered.connect(self.save_excel_as)
        file_menu.addAction(save_as_action)
        file_menu.addSeparator()
        self.recent_menu = QMenu("📁 最近打开的文件", file_menu)
        file_menu.addMenu(self.recent_menu)
        self.file_btn.setMenu(file_menu)
        top_layout.addWidget(self.file_btn)

        self.btn_export_csv = QPushButton(" 导出CSV")
        self.btn_export_csv.setIcon(qta.icon('fa5s.file-csv', color='#606266'))
        self.btn_export_csv.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_export_csv.setToolTip("将当前表格导出为 CSV 文件")
        self.btn_export_csv.clicked.connect(self.export_csv)
        top_layout.addWidget(self.btn_export_csv)

        self.sheet_combo = QComboBox()
        self.sheet_combo.setToolTip("切换工作表")
        self.sheet_combo.currentIndexChanged.connect(self.switch_sheet)
        top_layout.addWidget(self.sheet_combo)

        top_layout.addStretch()

        icon_color = "#333333" if not self.is_dark_mode else "#cccccc"
        toolbar_btn_style = """
            QPushButton {
                background-color: transparent; color: #409eff; border: 1px solid #d0e3ff;
                padding: 6px 12px; border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #409eff; color: white; border-color: #409eff;
            }
            QPushButton:pressed {
                background-color: #337ecc; color: white; border-color: #337ecc;
            }
        """
        self.btn_find = QPushButton(" 查找替换")
        self.btn_find.setIcon(qta.icon('fa5s.search', color=icon_color))
        self.btn_find.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_find.setToolTip("查找和替换单元格内容 (Ctrl+F)")
        self.btn_find.clicked.connect(self.show_find_replace)
        self.btn_find.setStyleSheet(toolbar_btn_style)
        top_layout.addWidget(self.btn_find)

        self.btn_freeze_row = QPushButton(" 冻结选中行")
        self.btn_freeze_row.setIcon(qta.icon('fa5s.snowflake', color=icon_color))
        self.btn_freeze_row.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_freeze_row.setToolTip("冻结当前选中的行，滚动时保持可见")
        self.btn_freeze_row.setCheckable(True)
        self.btn_freeze_row.clicked.connect(self.toggle_freeze_row)
        self.btn_freeze_row.setStyleSheet(toolbar_btn_style)
        top_layout.addWidget(self.btn_freeze_row)

        self.btn_zoom_in = QPushButton(" 放大")
        self.btn_zoom_in.setIcon(qta.icon('fa5s.search-plus', color=icon_color))
        self.btn_zoom_in.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_zoom_in.setToolTip("放大表格显示 (Ctrl+Plus)")
        self.btn_zoom_in.clicked.connect(self.zoom_in)
        self.btn_zoom_in.setStyleSheet(toolbar_btn_style)
        top_layout.addWidget(self.btn_zoom_in)

        self.btn_zoom_out = QPushButton(" 缩小")
        self.btn_zoom_out.setIcon(qta.icon('fa5s.search-minus', color=icon_color))
        self.btn_zoom_out.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_zoom_out.setToolTip("缩小表格显示 (Ctrl+Minus)")
        self.btn_zoom_out.clicked.connect(self.zoom_out)
        self.btn_zoom_out.setStyleSheet(toolbar_btn_style)
        top_layout.addWidget(self.btn_zoom_out)

        self.btn_zoom_reset = QPushButton(" 100%")
        self.btn_zoom_reset.setIcon(qta.icon('fa5s.sync-alt', color=icon_color))
        self.btn_zoom_reset.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_zoom_reset.setToolTip("重置缩放比例 (Ctrl+0)")
        self.btn_zoom_reset.clicked.connect(self.zoom_reset)
        self.btn_zoom_reset.setStyleSheet(toolbar_btn_style)
        top_layout.addWidget(self.btn_zoom_reset)

        layout.addLayout(top_layout)

        btn_layout = QHBoxLayout()
        self.btn_save = QPushButton(" 保存修改")
        self.btn_save.setIcon(qta.icon('fa5s.save', color='white'))
        self.btn_save.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_save.setToolTip("保存当前修改到原文件 (Ctrl+S)")
        self.btn_save.clicked.connect(self.save_excel)
        self.btn_save.setStyleSheet("""
            QPushButton {
                background-color: #10b981; color: white; border: none;
                padding: 6px 12px; border-radius: 6px; font-weight: 500;
            }
            QPushButton:hover { background-color: #059669; }
            QPushButton:disabled { background-color: #d1d5db; color: #9ca3af; }
        """)
        btn_layout.addWidget(self.btn_save)

        self.btn_add = QPushButton(" 添加")
        self.btn_add.setIcon(qta.icon('fa5s.plus', color='#409eff'))
        self.btn_add.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_add.setToolTip("添加行或列")
        self.btn_add.setStyleSheet("""
            QPushButton {
                background-color: #ffffff; color: #409eff; border: 1px solid #409eff;
                padding: 6px 12px; border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #ecf5ff;
            }
        """)
        add_menu = QMenu(self)
        add_row_action = QAction(" 添加行", self)
        add_row_action.setIcon(qta.icon('fa5s.plus', color='#409eff'))
        add_row_action.triggered.connect(self.add_row)
        add_menu.addAction(add_row_action)
        add_col_action = QAction(" 添加列", self)
        add_col_action.setIcon(qta.icon('fa5s.plus', color='#409eff'))
        add_col_action.triggered.connect(self.add_column)
        add_menu.addAction(add_col_action)
        self.btn_add.setMenu(add_menu)
        btn_layout.addWidget(self.btn_add)

        self.btn_delete = QPushButton(" 删除")
        self.btn_delete.setIcon(qta.icon('fa5s.trash-alt', color='#dc2626'))
        self.btn_delete.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_delete.setToolTip("删除选中的行或列")
        self.btn_delete.setStyleSheet("""
            QPushButton {
                background-color: #fef2f2; color: #dc2626; border: 1px solid #fecaca;
                padding: 6px 12px; border-radius: 6px;
            }
            QPushButton:hover { background-color: #fee2e2; }
        """)
        delete_menu = QMenu(self)
        delete_row_action = QAction(" 删除行", self)
        delete_row_action.setIcon(qta.icon('fa5s.trash-alt', color='#dc2626'))
        delete_row_action.triggered.connect(self.delete_row)
        delete_menu.addAction(delete_row_action)
        delete_col_action = QAction(" 删除列", self)
        delete_col_action.setIcon(qta.icon('fa5s.trash-alt', color='#dc2626'))
        delete_col_action.triggered.connect(self.delete_column)
        delete_menu.addAction(delete_col_action)
        self.btn_delete.setMenu(delete_menu)
        btn_layout.addWidget(self.btn_delete)

        self.btn_clean = QPushButton(" 清理空白行/列")
        self.btn_clean.setIcon(qta.icon('fa5s.broom', color='#409eff'))
        self.btn_clean.setIconSize(ICON_SIZE_MEDIUM)
        self.btn_clean.setToolTip("自动删除完全空白的行和列")
        self.btn_clean.clicked.connect(self.clean_empty_rows_columns)
        self.btn_clean.setStyleSheet("""
            QPushButton {
                background-color: #ffffff; color: #409eff; border: 1px solid #409eff;
                padding: 6px 12px; border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #ecf5ff;
            }
        """)
        btn_layout.addWidget(self.btn_clean)

        layout.addLayout(btn_layout)

        formula_layout = QHBoxLayout()
        self.cell_address_label = QLabel("A1")
        self.cell_address_label.setFixedWidth(50)
        formula_layout.addWidget(self.cell_address_label)
        self.label_fx = QLabel("fx")
        self.label_fx.setFixedWidth(30)
        formula_layout.addWidget(self.label_fx)
        self.formula_edit = QLineEdit()
        self.formula_edit.setPlaceholderText("输入公式或值")
        self.formula_edit.setToolTip("输入公式（以 = 开头）或直接输入值，按 Enter 确认")
        self.formula_edit.setContextMenuPolicy(Qt.CustomContextMenu)
        self.formula_edit.customContextMenuRequested.connect(self.show_formula_context_menu)
        formula_layout.addWidget(self.formula_edit)
        layout.addLayout(formula_layout)

        self.table = QTableWidget()
        self.table.setWordWrap(False)
        self.table.setAlternatingRowColors(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.itemChanged.connect(self.on_cell_changed)
        self.table.itemSelectionChanged.connect(self.update_selection_stats)

        header = CustomHeaderView(Qt.Horizontal, self.table)
        self.table.setHorizontalHeader(header)
        self.table_delegate = TableEditorDelegate()
        self.table.setItemDelegate(self.table_delegate)
        layout.addWidget(self.table)

        self.status_bar = QWidget()
        status_layout = QHBoxLayout(self.status_bar)
        status_layout.setContentsMargins(0, 0, 0, 0)
        self.sheet_label = QLabel("Sheet: -")
        status_layout.addWidget(self.sheet_label)
        status_layout.addStretch()
        self.cell_label = QLabel("单元格: -")
        status_layout.addWidget(self.cell_label)
        self.modified_label = QLabel("")
        status_layout.addWidget(self.modified_label)
        layout.addWidget(self.status_bar)

        self.table.currentItemChanged.connect(self.select_cell_change)
        self.table.cellClicked.connect(self.select_cell_change)
        self.table.doubleClicked.connect(self.on_cell_double_clicked)
        self.formula_edit.returnPressed.connect(self.confirm_formula_edit)
        self.table.verticalScrollBar().valueChanged.connect(self.on_vertical_scroll)
        self.create_shortcuts()

        if self.theme_manager:
            self.apply_theme()
        else:
            self.apply_table_style_compat()

        self._reset_to_initial_state()
        self._update_recent_menu()

        if not self.is_pro:
            self._set_readonly_mode()

    def apply_theme(self):
        if self.theme_manager:
            theme = self.theme_manager.get_theme()
            primary = theme["primary"]
            bg = theme["bg"]
            card = theme["card"]
            border = theme["border"]
            text = theme["text"]
            edit_bg = theme.get("edit_bg", card)
            
            dark_themes = ["dark", "pink", "dark_blue"]
            is_dark = self.theme_manager.current_theme in dark_themes

            if is_dark:
                edit_bg = "#252526"
                row_bg = "#252525"
                header_bg = "#2d2d2d"
                scroll_bg = "#1e1e1e"
                scroll_handle = "#4a4a4a"
                gridline = "#3a3a3a"
            else:
                edit_bg = "#ffffff"
                row_bg = "#f5f5f5"
                header_bg = "#f0f0f0"
                scroll_bg = "#f3f4f6"
                scroll_handle = "#9ca3af"
                gridline = "#d0d0d0"
            
            scrollbar_style = f"""
                QScrollBar:horizontal {{ height: 20px; background-color: {scroll_bg}; border-radius: 10px; margin: 0px 4px; }}
                QScrollBar:vertical {{ width: 20px; background-color: {scroll_bg}; border-radius: 10px; margin: 4px 0px; }}
                QScrollBar::handle:horizontal {{ background-color: {scroll_handle}; min-width: 40px; border-radius: 8px; }}
                QScrollBar::handle:vertical {{ background-color: {scroll_handle}; min-height: 40px; border-radius: 8px; }}
                QScrollBar::handle:hover {{ background-color: #6b7280; }}
                QScrollBar::add-line, QScrollBar::sub-line {{ width: 0px; height: 0px; }}
            """
            
            if is_dark:
                self.table.setStyleSheet("""
                    QTableWidget {
                        background-color: #1f1f1f;
                        color: #e0e0e0;
                        gridline-color: #3a3a3a;
                        border: 1px solid #3a3a3a;
                        selection-color: #ffffff;
                    }
                    QTableWidget::item {
                        padding: 0px 0px;
                        border: none;
                        border-right: 1px solid #3a3a3a;
                        border-bottom: 1px solid #3a3a3a;
                        background-color: #1f1f1f;
                        color: #e0e0e0;
                    }
                    QTableWidget::item:hover {
                        background-color: #2a2a2a;
                    }
                    QTableWidget::item:selected {
                        background-color: #409eff !important;
                        color: #ffffff !important;
                        border: 1px solid #66b1ff;
                    }
                    QTableWidget::item:selected:!active {
                        background-color: #2a2a2a !important;
                        color: #c0c0c0 !important;
                    }
                    QTableWidget::item:edit {
                        background-color: #1f1f1f;
                        color: #e0e0e0;
                        border: 2px solid #66b1ff;
                        border-radius: 0px;
                        padding: 0px;
                    }
                    QTableWidget::item:selected:edit {
                        background-color: #1f1f1f;
                        color: #e0e0e0;
                        border: 2px solid #66b1ff;
                    }
                    QTableWidget QLineEdit {
                        background-color: transparent !important;
                        color: #e0e0e0 !important;
                        border: none !important;
                        border-radius: 0px !important;
                        padding: 0px !important;
                        margin: 0px !important;
                        font-size: 13px !important;
                        min-height: 100% !important;
                        width: 100% !important;
                        selection-background-color: #66b1ff;
                        selection-color: white;
                    }
                    QHeaderView::section {
                        background-color: #2d2d2d;
                        color: #c0c0c0;
                        border: 1px solid #3a3a3a;
                        border-top: none;
                        border-left: none;
                        padding: 6px 8px;
                        font-weight: 500;
                        font-size: 12px;
                    }
                    QHeaderView::section:hover {
                        background-color: #3a3a3a;
                    }
                    QHeaderView::section:pressed {
                        background-color: #4a4a4a;
                    }
                    QTableWidget::verticalHeader {
                        background-color: #2d2d2d;
                        border: 1px solid #3a3a3a;
                        border-left: none;
                        border-top: none;
                    }
                    QTableWidget::verticalHeader::section {
                        background-color: #2d2d2d;
                        color: #c0c0c0;
                        border: 1px solid #3a3a3a;
                        border-left: none;
                        border-top: none;
                        padding: 6px 8px;
                        font-size: 12px;
                    }
                    QTableWidget::verticalHeader::section:hover {
                        background-color: #3a3a3a;
                    }
                    QTableWidget QLineEdit:focus {
                        background-color: transparent !important;
                        border: none !important;
                        outline: none !important;
                    }
                """ + scrollbar_style)
            else:
                self.table.setStyleSheet("""
                    QTableWidget {
                        background-color: white;
                        color: #303133;
                        gridline-color: #e4e7ed;
                        border: 1px solid #e4e7ed;
                        selection-color: #303133;
                    }
                    QTableWidget::item {
                        padding: 0px 0px;
                        border: none;
                        border-right: 1px solid #f0f0f0;
                        border-bottom: 1px solid #f0f0f0;
                    }
                    QTableWidget::item:hover {
                        background-color: #f5f7fa;
                    }
                    QTableWidget::item:selected {
                        background-color: rgba(64, 158, 255, 0.2);
                        color: #303133;
                        border: 1px solid rgba(64, 158, 255, 0.3);
                    }
                    QTableWidget::item:selected:!active {
                        background-color: #f5f7fa;
                        color: #606266;
                    }
                    QTableWidget::item:edit {
                        background-color: white;
                        color: #303133;
                        selection-background-color: white;
                        selection-color: #303133;
                        border: 2px solid rgba(64, 158, 255, 0.5);
                        border-radius: 0px;
                        padding: 0px;
                    }
                    QTableWidget::item:selected:edit {
                        background-color: white;
                        color: #303133;
                        border: 2px solid rgba(64, 158, 255, 0.5);
                    }
                    QTableWidget QLineEdit {
                        background-color: transparent !important;
                        color: #303133 !important;
                        border: none !important;
                        border-radius: 0px !important;
                        padding: 0px !important;
                        margin: 0px !important;
                        font-size: 13px !important;
                        min-height: 100% !important;
                        width: 100% !important;
                        selection-background-color: #409eff;
                        selection-color: white;
                    }                    
                    QHeaderView::section {
                        background-color: #f0f0f0;
                        color: #303133;
                        border: 1px solid #e4e7ed;
                        border-top: none;
                        border-left: none;
                        padding: 6px 8px;
                        font-weight: 500;
                        font-size: 12px;
                    }
                    QHeaderView::section:hover {
                        background-color: #e4e7ed;
                    }
                    QHeaderView::section:pressed {
                        background-color: #d9d9d9;
                    }
                    QTableWidget::verticalHeader {
                        background-color: #f0f0f0;
                        border: 1px solid #e4e7ed;
                        border-left: none;
                        border-top: none;
                    }
                    QTableWidget::verticalHeader::section {
                        background-color: #f0f0f0;
                        color: #303133;
                        border: 1px solid #e4e7ed;
                        border-left: none;
                        border-top: none;
                        padding: 6px 8px;
                        font-size: 12px;
                    }
                    QTableWidget::verticalHeader::section:hover {
                        background-color: #e4e7ed;
                    }
                    QTableWidget QLineEdit:focus {
                        background-color: transparent !important;
                        border: none !important;
                        outline: none !important;
                    }
                """ + scrollbar_style)
            
            self.formula_edit.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {edit_bg};
                    border: 1px solid {border};
                    border-radius: 4px;
                    padding: 4px 8px;
                    color: {text};
                    font-family: Consolas, monospace;
                }}
                QLineEdit:focus {{ border-color: {primary}; }}
            """)
            
            if hasattr(self, 'status_bar'):
                self.status_bar.setStyleSheet(f"""
                    QWidget {{ background-color: {bg}; }}
                    QLabel {{ color: {text}; }}
                """)
            
            # 不再刷新样式（已空）
        else:
            self.apply_table_style_compat()

    def apply_table_style_compat(self):
        if self.is_dark_mode:
            scrollbar_style = """
                QScrollBar:horizontal { height: 20px; background-color: #1e1e1e; border-radius: 10px; margin: 0px 4px; }
                QScrollBar:vertical { width: 20px; background-color: #1e1e1e; border-radius: 10px; margin: 4px 0px; }
                QScrollBar::handle:horizontal { background-color: #4a4a4a; min-width: 40px; border-radius: 8px; }
                QScrollBar::handle:vertical { background-color: #4a4a4a; min-height: 40px; border-radius: 8px; }
                QScrollBar::handle:hover { background-color: #6b7280; }
                QScrollBar::add-line, QScrollBar::sub-line { width: 0px; height: 0px; }
            """
            self.table.setStyleSheet("""
                QTableWidget { 
                    background-color: #252526; 
                    color: #c0c0c0; 
                    gridline-color: #3a3a3a;
                    border: 1px solid #3a3a3a;
                    selection-color: #c0c0c0;
                }
                QTableWidget::item { 
                    padding: 4px;
                    border: none;
                    background-color: #252526;
                }
                QTableWidget::item:hover { background-color: #323232; }
                QTableWidget::item:selected { background-color: rgba(64, 158, 255, 0.25); color: #e0e0e0; }
                QTableWidget::item:selected:!active { background-color: #323232; color: #c0c0c0; }
                QTableWidget::item:edit { 
                    background-color: #252526; 
                    color: #e0e0e0; 
                    selection-background-color: #252526;
                    selection-color: #e0e0e0;
                    border: none;
                    margin: 0;
                    padding: 0;
                }
                QTableWidget::item:selected:edit { 
                    background-color: #252526; 
                    color: #e0e0e0; 
                }
                QHeaderView::section { 
                    background-color: #2d2d2d; 
                    color: #c0c0c0; 
                    border: 1px solid #3a3a3a;
                    border-top: none;
                    border-left: none;
                    padding: 6px;
                    font-weight: 500;
                }
                QHeaderView::section:hover { background-color: #3a3a3a; }
                QTableWidget::verticalHeader { 
                    background-color: #2d2d2d;
                    border: 1px solid #3a3a3a;
                    border-left: none;
                    border-top: none;
                }
                QTableWidget::verticalHeader::section {
                    background-color: #2d2d2d;
                    color: #c0c0c0;
                    border: 1px solid #3a3a3a;
                    border-left: none;
                    border-top: none;
                    padding: 6px;
                }
                QTableWidget QLineEdit {
                    background-color: #252526;
                    color: #e0e0e0;
                    border: none;
                    border-radius: 0;
                    padding: 4px 8px;
                    margin: 0;
                    font-size: 13px;
                    selection-background-color: #409eff;
                    selection-color: white;
                }
                QTableWidget QLineEdit:focus {
                    background-color: #252526;
                    border: none;
                    outline: none;
                }
            """ + scrollbar_style)
            self.formula_edit.setStyleSheet("""
                QLineEdit {
                    background-color: #1a1a1a;
                    border: 1px solid #3a3a3a;
                    border-radius: 4px;
                    padding: 4px 8px;
                    color: #e0e0e0;
                    font-family: Consolas, monospace;
                }
                QLineEdit:focus { border-color: #409eff; }
            """)
        else:
            scrollbar_style = """
                QScrollBar:horizontal { height: 20px; background-color: #f3f4f6; border-radius: 10px; margin: 0px 4px; }
                QScrollBar:vertical { width: 20px; background-color: #f3f4f6; border-radius: 10px; margin: 4px 0px; }
                QScrollBar::handle:horizontal { background-color: #9ca3af; min-width: 40px; border-radius: 8px; }
                QScrollBar::handle:vertical { background-color: #9ca3af; min-height: 40px; border-radius: 8px; }
                QScrollBar::handle:hover { background-color: #6b7280; }
                QScrollBar::add-line, QScrollBar::sub-line { width: 0px; height: 0px; }
            """
            self.table.setStyleSheet("""
                QTableWidget { background-color: #ffffff; alternate-background-color: #f5f5f5; color: #000000; gridline-color: #d0d0d0; }
                QTableWidget::item:selected { background-color: #409eff; color: #ffffff; }
                QHeaderView::section { background-color: #f0f0f0; color: #000000; border: 1px solid #d0d0d0; }
            """ + scrollbar_style)
            self.formula_edit.setStyleSheet("""
                QLineEdit {
                    background-color: #ffffff;
                    border: 1px solid #dcdfe6;
                    border-radius: 4px;
                    padding: 4px 8px;
                    color: #303133;
                    font-family: Consolas, monospace;
                }
                QLineEdit:focus { border-color: #409eff; }
            """)

    def _reset_to_initial_state(self):
        self._create_default_blank_table()
        self.cell_formula.clear()
        self.cell_result.clear()
        self.cell_styles.clear()
        self.merged_cells = []
        self.original_workbook = None
        self.file_path = None
        self.modified = False
        self.modified_label.setText("")
        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItem("Sheet1")
        self.sheet_combo.setEnabled(True)
        self.sheet_combo.blockSignals(False)
        self.btn_save.setEnabled(self.is_pro)
        self.btn_add.setEnabled(self.is_pro)
        self.btn_delete.setEnabled(self.is_pro)
        self.btn_clean.setEnabled(self.is_pro)
        self.btn_export_csv.setEnabled(self.is_pro)
        self.btn_find.setEnabled(self.is_pro)
        self.btn_freeze_row.setEnabled(self.is_pro)
        self.btn_zoom_in.setEnabled(True)
        self.btn_zoom_out.setEnabled(True)
        self.btn_zoom_reset.setEnabled(True)
        self.btn_freeze_row.setChecked(False)
        self.btn_freeze_row.setText(" 冻结选中行")
        self.sheet_label.setText("Sheet: Sheet1")
        self.cell_address_label.setText("A1")
        self.formula_edit.clear()
        self.setWindowTitle("Excel数据编辑" + (" - 新建文件" if self.is_pro else "（只读模式 - 免费版）"))
        
        try:
            fd, self.backup_file = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
        except:
            self.backup_file = None

    def _create_default_blank_table(self):
        rows = 50
        cols = 20
        self.table.setRowCount(rows)
        self.table.setColumnCount(cols)

        header_labels = []
        for i in range(cols):
            if i < 26:
                header_labels.append(chr(65 + i))
            else:
                first = chr(65 + (i // 26) - 1)
                second = chr(65 + (i % 26))
                header_labels.append(f"{first}{second}")
        self.table.setHorizontalHeaderLabels(header_labels)

        for i in range(rows):
            self.table.setRowHeight(i, 25)
        for i in range(cols):
            self.table.setColumnWidth(i, 100)

        for r in range(rows):
            for c in range(cols):
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(r, c, item)
        # 不设置任何背景，样式表自动生效
        
        # 新增：从第一个单元格读取基准字体
        item = self.table.item(0, 0)
        if item:
            self._base_font_size = item.font().pointSize() or 10
        # 基准行高和列宽保持不变
        self._base_row_height = 25
        self._base_col_width = 100

    def _save_original_fonts(self):
        self.original_font_sizes.clear()
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                if item:
                    self.original_font_sizes[(r, c)] = item.font().pointSize()

    def _apply_zoom(self):
        """应用缩放：同时调整字体、行高、列宽"""
        # 确保基准值存在
        if not hasattr(self, '_base_font_size') or self._base_font_size <= 0:
            self._base_font_size = 10
        if not hasattr(self, '_base_row_height') or self._base_row_height <= 0:
            self._base_row_height = 25
        if not hasattr(self, '_base_col_width') or self._base_col_width <= 0:
            self._base_col_width = 100
        
        # 计算新尺寸
        new_font_size = max(6, int(self._base_font_size * self.zoom_factor))
        new_row_height = max(15, int(self._base_row_height * self.zoom_factor))
        new_col_width = max(30, int(self._base_col_width * self.zoom_factor))
        
        # 设置字体大小
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                if item:
                    font = item.font()
                    font.setPointSize(new_font_size)
                    item.setFont(font)
            # 设置行高
            self.table.setRowHeight(r, new_row_height)
        
        # 设置列宽
        for c in range(self.table.columnCount()):
            self.table.setColumnWidth(c, new_col_width)
        
        # 更新公式栏字体
        formula_font = self.formula_edit.font()
        formula_font.setPointSize(max(10, new_font_size))
        self.formula_edit.setFont(formula_font)
        
        # 更新按钮文字
        self.btn_zoom_reset.setText(f"{int(self.zoom_factor * 100)}%")

    def zoom_in(self):
        if not self._check_pro_edit_permission():
            return
        self.zoom_factor = min(2.0, self.zoom_factor + 0.1)
        self._apply_zoom()

    def zoom_out(self):
        if not self._check_pro_edit_permission():
            return
        self.zoom_factor = max(0.5, self.zoom_factor - 0.1)
        self._apply_zoom()

    def zoom_reset(self):
        if not self._check_pro_edit_permission():
            return
        self.zoom_factor = 1.0
        # 重置基准
        self._base_font_size = 10
        self._base_row_height = 25
        self._base_col_width = 100
        # 从原始记录中恢复字体基准（如果有）
        if self.original_font_sizes and (0, 0) in self.original_font_sizes:
            self._base_font_size = self.original_font_sizes[(0, 0)]
        self._apply_zoom()

    def toggle_freeze_row(self):
        if not self._check_pro_edit_permission():
            self.btn_freeze_row.setChecked(False)
            return
        if self.btn_freeze_row.isChecked():
            cur_row = self.table.currentRow()
            if cur_row < 0:
                QMessageBox.warning(self, "提示", "请先选中要冻结的行")
                self.btn_freeze_row.setChecked(False)
                return
            self.frozen_row = cur_row
            self.btn_freeze_row.setText(f" 已冻结第{cur_row+1}行")
            self.on_vertical_scroll(self.table.verticalScrollBar().value())
        else:
            self.frozen_row = -1
            self.btn_freeze_row.setText(" 冻结选中行")

    def on_vertical_scroll(self, value):
        if self.frozen_row >= 0 and self.frozen_row < self.table.rowCount():
            item = self.table.item(self.frozen_row, 0)
            if item:
                self.table.scrollToItem(item, QTableWidget.PositionAtTop)

    def show_formula_context_menu(self, pos):
        if not self.is_pro:
            return
        menu = QMenu(self)
        if self.theme_manager:
            theme = self.theme_manager.get_theme()
            menu.setStyleSheet(f"QMenu {{ background-color: {theme['card']}; border: 1px solid {theme['border']}; border-radius: 8px; color: {theme['text']}; }}")
        copy_action = QAction(" 复制", self)
        copy_action.setIcon(qta.icon('fa5s.copy', color='#606266'))
        copy_action.setToolTip("复制选中文本")
        copy_action.triggered.connect(self.formula_edit.copy)
        menu.addAction(copy_action)
        paste_action = QAction(" 粘贴", self)
        paste_action.setIcon(qta.icon('fa5s.paste', color='#606266'))
        paste_action.setToolTip("粘贴文本")
        paste_action.triggered.connect(self.formula_edit.paste)
        menu.addAction(paste_action)
        menu.addSeparator()
        sym_sub = menu.addMenu("📝 符号处理（应用于选中文本）")
        cfg = self.get_symbol_config()
        
        default_syms = {}
        custom_syms = {}
        for sk, si in cfg.items():
            if sk.startswith("custom_"):
                custom_syms[sk] = si
            else:
                default_syms[sk] = si
        
        default_sub = sym_sub.addMenu("默认符号")
        for sk, si in default_syms.items():
            menu_title = f"{si[0]}{si[1]}{si[3]}"
            sm = QMenu(menu_title, default_sub)
            for pl, pk in [("头部","head"),("尾部","tail"),("两端","both"),("分隔","split")]:
                a = QAction(pl, sm)
                a.triggered.connect(lambda checked, s=sk, p=pk: self.apply_symbol_to_selected_formula_text(s, p))
                sm.addAction(a)
            default_sub.addMenu(sm)
        
        if custom_syms:
            custom_menu = sym_sub.addMenu("⭐ 自定义符号")
            for sk, si in custom_syms.items():
                menu_title = f"{si[0]}{si[1]}{si[3]}"
                sm = QMenu(menu_title, custom_menu)
                for pl, pk in [("头部","head"),("尾部","tail"),("两端","both"),("分隔","split")]:
                    a = QAction(pl, sm)
                    a.triggered.connect(lambda checked, s=sk, p=pk: self.apply_symbol_to_selected_formula_text(s, p))
                    sm.addAction(a)
                custom_menu.addMenu(sm)
        
        menu.exec_(self.formula_edit.mapToGlobal(pos))

    def apply_symbol_to_selected_formula_text(self, sym, pos):
        if not self.is_pro:
            return
        selected_text = self.formula_edit.selectedText()
        if not selected_text:
            QMessageBox.warning(self, "提示", "请先在公式栏选中要处理的文本")
            return
        cfg = self.get_symbol_config()
        if sym not in cfg:
            return
        left, right = cfg[sym][0], cfg[sym][1] if len(cfg[sym]) >= 2 else cfg[sym][0]
        processed = self._apply_symbol_to_text(selected_text, left, right, pos, self.split_mode)
        selection_start = self.formula_edit.selectionStart()
        selection_end = selection_start + len(selected_text)
        text = self.formula_edit.text()
        new_text = text[:selection_start] + processed + text[selection_end:]
        self.formula_edit.setText(new_text)
        self.formula_edit.setCursorPosition(selection_start + len(processed))

    def _apply_symbol_to_text(self, text, left, right, position, split_mode):
        pure = text.strip()
        if not pure:
            return text
        if position == "head":
            return left + pure
        elif position == "tail":
            return pure + right
        elif position == "both":
            return left + pure + right
        elif position == "split":
            chars = list(pure)
            if split_mode == 0:
                return left.join(chars)
            elif split_mode == 1:
                return "".join([f"{left}{c}{right}" for c in chars])
            elif split_mode == 2:
                return ",".join([f"{left}{c}{right}" for c in chars])
        return pure

    def import_excel(self):
        if not self.is_pro:
            QMessageBox.warning(self, "权限不足", "打开文件为 PRO 版专属")
            return
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)")
        if not path:
            return
        ok, err = self._load_excel_data(path)
        if ok:
            self._show_function_support_tip()
            QMessageBox.information(self, "成功", f"数据导入成功！\n共 {len(self.sheet_names)} 个工作表")
        else:
            QMessageBox.critical(self, "错误", f"导入失败：{err}")

    def _open_recent_file(self, file_path):
        if not self.is_pro:
            return
        if not os.path.exists(file_path):
            QMessageBox.warning(self, "提示", f"文件不存在：{file_path}")
            if file_path in self.recent_files:
                self.recent_files.remove(file_path)
            return
        ok, err = self._load_excel_data(file_path)
        if ok:
            self.setWindowTitle(f"Excel数据编辑 - {os.path.basename(file_path)}")
        else:
            QMessageBox.critical(self, "错误", f"打开文件失败：{err}")

    def open_file_direct(self, file_path):
        if not self.is_pro:
            return
        self._open_recent_file(file_path)

    def switch_sheet(self, index):
        if not self.is_pro:
            return
        if self.file_path is None or index < 0 or index >= len(self.sheet_names):
            return
        if self.modified and self.is_pro:
            reply = QMessageBox.question(self, "提示", "当前工作表有未保存的修改，是否保存？",
                                         QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
            if reply == QMessageBox.Save:
                self.save_excel()
            elif reply == QMessageBox.Cancel:
                self.sheet_combo.setCurrentIndex(self.current_sheet)
                return
        sheet_name = self.sheet_names[index]
        ok, err = self._load_excel_data(self.file_path, sheet_name)
        if not ok:
            QMessageBox.critical(self, "错误", f"加载工作表失败：{err}")
            self.sheet_combo.setCurrentIndex(self.current_sheet)
        else:
            self.current_sheet = index

    def _add_to_recent_files(self, file_path):
        if not self.is_pro:
            return
        if file_path in self.recent_files:
            self.recent_files.remove(file_path)
        self.recent_files.insert(0, file_path)
        if len(self.recent_files) > 10:
            self.recent_files.pop()
        if self.main_window and hasattr(self.main_window, 'add_to_recent_files'):
            self.main_window.add_to_recent_files(file_path)

    def save_excel(self, custom_path=None):
        if custom_path is not None and not isinstance(custom_path, str):
            print(f"无效的 custom_path 类型: {type(custom_path)}")
            QMessageBox.critical(self, "保存错误", "保存路径无效，请使用「另存为」重新选择。")
            self.file_path = None
            self.save_excel_as()
            return
        if not self._check_pro_edit_permission():
            return
        save_path = custom_path if custom_path is not None else self.file_path
        if save_path is None:
            self.save_excel_as()
            return
        save_dir = os.path.dirname(save_path)
        if save_dir and not os.path.exists(save_dir):
            QMessageBox.critical(self, "路径错误", f"保存目录不存在：{save_dir}\n请使用「另存为」重新选择。")
            self.save_excel_as()
            return
        try:
            from openpyxl import load_workbook, Workbook
            if os.path.exists(save_path):
                wb = load_workbook(save_path, data_only=False)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
            merged_topleft = set()
            all_merged = set()
            for merged_range in ws.merged_cells.ranges:
                topleft = (merged_range.min_row, merged_range.min_col)
                merged_topleft.add(topleft)
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        all_merged.add((row, col))
            rows = self.table.rowCount()
            cols = self.table.columnCount()
            for r in range(rows):
                for c in range(cols):
                    excel_row = r + 1
                    excel_col = c + 1
                    if (excel_row, excel_col) in all_merged and (excel_row, excel_col) not in merged_topleft:
                        continue
                    formula = self.cell_formula.get((r, c), "")
                    val = self.cell_result.get((r, c), "")
                    cell = ws.cell(row=excel_row, column=excel_col)
                    if formula and formula.startswith("="):
                        cell.value = formula
                    else:
                        cell.value = val if val else ""
                    if (r, c) in self.cell_styles:
                        from copy import copy
                        style = self.cell_styles[(r, c)]
                        if style['font']:
                            cell.font = copy(style['font'])
                        if style['fill']:
                            cell.fill = copy(style['fill'])
                        if style['border']:
                            cell.border = copy(style['border'])
                        if style['alignment']:
                            cell.alignment = copy(style['alignment'])
                        if style['number_format']:
                            cell.number_format = style['number_format']
            wb.save(save_path)
            if custom_path is not None:
                self.file_path = save_path
            self.modified = False
            self.modified_label.setText("")
            self._clear_all_highlights()
            self.edited_cells.clear()
            QMessageBox.information(self, "成功", "文件已保存！")
        except PermissionError as e:
            logger.error(f"保存Excel失败（权限错误）: {e}", exc_info=True)
            QMessageBox.critical(self, "错误", f"保存失败：文件被占用或权限不足。\n请关闭 Excel 中打开的文件后重试。\n详细信息：{str(e)}")
        except Exception as e:
            logger.error(f"保存Excel失败: {e}", exc_info=True)
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}\n\n建议：尝试使用「另存为」保存到新文件。")

    def save_excel_as(self):
        if not self._check_pro_edit_permission():
            return
        if self.table.rowCount() == 0 or self.table.columnCount() == 0:
            QMessageBox.warning(self, "提示", "没有可保存的数据")
            return
        default_name = "未命名.xlsx"
        if self.file_path:
            default_name = os.path.basename(self.file_path)
        from pathlib import Path
        default_dir = str(Path.home() / "Desktop")
        if not os.path.exists(default_dir):
            default_dir = str(Path.home() / "Downloads")
        if not os.path.exists(default_dir):
            default_dir = str(Path.home())
        default_path = os.path.join(default_dir, default_name)
        while True:
            save_path, _ = QFileDialog.getSaveFileName(self, "另存为", default_path, "Excel文件 (*.xlsx)")
            if not save_path:
                return
            if not save_path.endswith('.xlsx'):
                save_path += '.xlsx'
            save_dir = os.path.dirname(save_path)
            if not os.path.exists(save_dir):
                QMessageBox.critical(self, "路径错误", f"保存目录不存在：{save_dir}\n请重新选择。")
                default_path = os.path.join(default_dir, default_name)
                continue
            try:
                self.save_excel(custom_path=save_path)
                self.file_path = save_path
                self.setWindowTitle(f"Excel数据编辑 - {os.path.basename(save_path)}")
                break
            except Exception as e:
                logger.error(f"另存为Excel失败: {e}", exc_info=True)
                QMessageBox.critical(self, "错误", f"保存失败：{str(e)}\n请重新选择保存位置。")
                default_path = save_path
                continue

    def _update_recent_menu(self):
        if not self.is_pro:
            return
        self.recent_menu.clear()
        if self.recent_files:
            for fp in self.recent_files[:10]:
                if os.path.exists(fp):
                    action = QAction(os.path.basename(fp), self)
                    action.triggered.connect(lambda checked, path=fp: self._open_recent_file(path))
                    self.recent_menu.addAction(action)
        else:
            action = QAction("暂无最近文件", self)
            action.setEnabled(False)
            self.recent_menu.addAction(action)

    def export_csv(self):
        if not self._check_pro_edit_permission():
            return
        path, _ = QFileDialog.getSaveFileName(self, "导出 CSV", "", "CSV文件 (*.csv)")
        if not path:
            return
        try:
            import csv
            rows = self.table.rowCount()
            cols = self.table.columnCount()
            
            def clean_text(text):
                """清理文本，确保CSV格式正确"""
                if text is None:
                    return ""
                # 替换可能导致CSV格式错误的特殊字符
                text = str(text)
                # csv.writer会自动处理包含换行、逗号、引号的文本
                # 但我们可以做一些基础清理
                return text
            
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
                header = [clean_text(self.table.horizontalHeaderItem(c).text() if self.table.horizontalHeaderItem(c) else f"Col{c+1}") for c in range(cols)]
                writer.writerow(header)
                for r in range(rows):
                    row_data = [clean_text(self.table.item(r, c).text() if self.table.item(r, c) else "") for c in range(cols)]
                    writer.writerow(row_data)
            QMessageBox.information(self, "成功", f"已导出到：{path}")
        except Exception as e:
            logger.error(f"导出CSV失败: {e}", exc_info=True)
            QMessageBox.critical(self, "错误", f"导出失败：{str(e)}")

    def show_find_replace(self):
        if not self._check_pro_edit_permission():
            return
        dlg = FindReplaceDialog(self)
        dlg.exec()

    def update_selection_stats(self):
        selected = self.table.selectedItems()
        if not selected:
            self.cell_label.setText("单元格: -")
            return
        selected_rows = set()
        selected_cols = set()
        for item in selected:
            selected_rows.add(item.row())
            selected_cols.add(item.column())
        row_count = len(selected_rows)
        col_count = len(selected_cols)
        cell_count = len(selected)
        numbers = []
        for item in selected:
            try:
                num = float(item.text())
                numbers.append(num)
            except:
                pass
        if row_count == 1 and col_count == 1:
            stats = f"单元格: {chr(65 + selected[0].column())}{selected[0].row() + 1}"
        elif cell_count == row_count:
            stats = f"行: {row_count}"
        elif cell_count == col_count:
            stats = f"列: {col_count}"
        else:
            stats = f"单元格: {cell_count}"
        if numbers:
            total = sum(numbers)
            avg = total / len(numbers)
            stats += f"  求和: {total:.2f}  平均值: {avg:.2f}"
        self.cell_label.setText(stats)

    def merge_selected_cells(self):
        if not self._check_pro_edit_permission():
            return
        selected = self.table.selectedRanges()
        if not selected:
            QMessageBox.warning(self, "提示", "请先选择要合并的单元格区域")
            return
        for sel in selected:
            top = sel.topRow()
            left = sel.leftColumn()
            bottom = sel.bottomRow()
            right = sel.rightColumn()
            if top == bottom and left == right:
                continue
            
            # 收集合并前的单元格内容，用于撤销
            cell_contents = {}
            for r in range(top, bottom+1):
                for c in range(left, right+1):
                    item = self.table.item(r, c)
                    cell_contents[(r, c)] = item.text() if item else ""
            
            # 记录撤销动作
            self.push_undo_action(None, None, None, None, {
                'type': 'merge',
                'top': top,
                'left': left,
                'bottom': bottom,
                'right': right,
                'cell_contents': cell_contents,
                'old_merged': list(self.merged_cells)
            })
            
            for r in range(top, bottom+1):
                for c in range(left, right+1):
                    self.table.setSpan(r, c, 1, 1)
            new_merged = []
            for (mr1, mc1, mr2, mc2) in self.merged_cells:
                if not (max(mr1, top) <= min(mr2, bottom) and max(mc1, left) <= min(mc2, right)):
                    new_merged.append((mr1, mc1, mr2, mc2))
            self.merged_cells = new_merged
            self.table.setSpan(top, left, bottom-top+1, right-left+1)
            main_item = self.table.item(top, left)
            main_text = main_item.text() if main_item else ""
            for r in range(top, bottom+1):
                for c in range(left, right+1):
                    if r == top and c == left:
                        continue
                    item = self.table.item(r, c)
                    if item:
                        item.setText("")
            self.merged_cells.append((top, left, bottom, right))
            self.modified = True
            self.modified_label.setText("*")
        QMessageBox.information(self, "成功", "已合并所选单元格")

    def split_selected_cell(self):
        if not self._check_pro_edit_permission():
            return
        current = self.table.currentItem()
        if not current:
            return
        r, c = current.row(), current.column()
        row_span = self.table.rowSpan(r, c)
        col_span = self.table.columnSpan(r, c)
        
        # 找到完整的合并区域
        top, left, bottom, right = r, c, r + row_span - 1, c + col_span - 1
        found = None
        for (mr1, mc1, mr2, mc2) in self.merged_cells:
            if mr1 <= r <= mr2 and mc1 <= c <= mc2:
                found = (mr1, mc1, mr2, mc2)
                top, left, bottom, right = mr1, mc1, mr2, mc2
                break
        
        if row_span == 1 and col_span == 1 and not found:
            QMessageBox.warning(self, "提示", "当前单元格未被合并")
            return
        
        # 收集合并前的信息用于撤销
        main_item = self.table.item(top, left)
        main_text = main_item.text() if main_item else ""
        
        # 记录撤销动作
        self.push_undo_action(None, None, None, None, {
            'type': 'split',
            'top': top,
            'left': left,
            'bottom': bottom,
            'right': right,
            'main_text': main_text,
            'old_merged': list(self.merged_cells)
        })
        
        self.table.setSpan(r, c, 1, 1)
        new_merged = []
        for (mr1, mc1, mr2, mc2) in self.merged_cells:
            if not (mr1 <= r <= mr2 and mc1 <= c <= mc2):
                new_merged.append((mr1, mc1, mr2, mc2))
        self.merged_cells = new_merged
        self.modified = True
        self.modified_label.setText("*")
        QMessageBox.information(self, "成功", "已拆分单元格")

    def add_row(self):
        if not self._check_pro_edit_permission():
            return
        if self.table.columnCount() == 0:
            QMessageBox.warning(self, "提示", "请先导入Excel文件！")
            return
        num = 1
        sel_rows = set(item.row() for item in self.table.selectedItems())
        if len(sel_rows) > 1:
            num, ok = QInputDialog.getInt(self, "批量添加行", "请输入要添加的行数:", 1, 1, 100, 1)
            if not ok:
                return
        for _ in range(num):
            row = self.table.rowCount()
            self.table.insertRow(row)
            for c in range(self.table.columnCount()):
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, c, item)
                self.cell_result[(row, c)] = ""
                self.cell_formula[(row, c)] = ""
        self.modified = True
        self.modified_label.setText("*")
        self.table.scrollToBottom()
        if num > 1:
            QMessageBox.information(self, "成功", f"已添加 {num} 行！")

    def add_column(self):
        if not self._check_pro_edit_permission():
            return
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "请先导入Excel文件！")
            return
        num = 1
        sel_cols = set(item.column() for item in self.table.selectedItems())
        if len(sel_cols) > 1:
            num, ok = QInputDialog.getInt(self, "批量添加列", "请输入要添加的列数:", 1, 1, 100, 1)
            if not ok:
                return
        for _ in range(num):
            col = self.table.columnCount()
            self.table.insertColumn(col)
            col_label = ""
            if col < 26:
                col_label = chr(65 + col)
            else:
                first = chr(65 + (col // 26) - 1)
                second = chr(65 + (col % 26))
                col_label = f"{first}{second}"
            self.table.setHorizontalHeaderItem(col, QTableWidgetItem(col_label))
            for r in range(self.table.rowCount()):
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(r, col, item)
                self.cell_result[(r, col)] = ""
                self.cell_formula[(r, col)] = ""
        self.modified = True
        self.modified_label.setText("*")
        if num > 1:
            QMessageBox.information(self, "成功", f"已添加 {num} 列！")

    def delete_row(self):
        if not self._check_pro_edit_permission():
            return
        sel_rows = set(item.row() for item in self.table.selectedItems())
        if not sel_rows:
            QMessageBox.warning(self, "提示", "请先选择要删除的行！")
            return
        reply = QMessageBox.question(self, "确认", f"确定要删除选中的 {len(sel_rows)} 行吗？")
        if reply == QMessageBox.Yes:
            for row in sorted(sel_rows, reverse=True):
                self.table.removeRow(row)
                for c in range(self.table.columnCount()):
                    self.cell_result.pop((row, c), None)
                    self.cell_formula.pop((row, c), None)
                    self.cell_styles.pop((row, c), None)
            self.modified = True
            self.modified_label.setText("*")

    def delete_column(self):
        if not self._check_pro_edit_permission():
            return
        sel_cols = set(item.column() for item in self.table.selectedItems())
        if not sel_cols:
            QMessageBox.warning(self, "提示", "请先选择要删除的列！")
            return
        reply = QMessageBox.question(self, "确认", f"确定要删除选中的 {len(sel_cols)} 列吗？")
        if reply == QMessageBox.Yes:
            for col in sorted(sel_cols, reverse=True):
                self.table.removeColumn(col)
                for r in range(self.table.rowCount()):
                    self.cell_result.pop((r, col), None)
                    self.cell_formula.pop((r, col), None)
                    self.cell_styles.pop((r, col), None)
            self.modified = True
            self.modified_label.setText("*")

    def insert_row_below(self):
        if not self._check_pro_edit_permission():
            return
        cur_row = self.table.currentRow()
        if cur_row < 0:
            cur_row = self.table.rowCount() - 1 if self.table.rowCount() > 0 else -1
        if cur_row < 0:
            self.add_row()
        else:
            self.table.insertRow(cur_row + 1)
            for c in range(self.table.columnCount()):
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(cur_row + 1, c, item)
                self.cell_result[(cur_row+1, c)] = ""
                self.cell_formula[(cur_row+1, c)] = ""
            self.modified = True
            self.modified_label.setText("*")

    def clean_empty_rows_columns(self):
        if not self._check_pro_edit_permission():
            return
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        if rows == 0 or cols == 0:
            QMessageBox.information(self, "提示", "表格为空，无需清理！")
            return
        empty_rows = []
        for r in range(rows):
            if all(not (self.table.item(r, c) and self.table.item(r, c).text().strip()) for c in range(cols)):
                empty_rows.append(r)
        empty_cols = []
        for c in range(cols):
            if all(not (self.table.item(r, c) and self.table.item(r, c).text().strip()) for r in range(rows)):
                empty_cols.append(c)
        if not empty_rows and not empty_cols:
            QMessageBox.information(self, "提示", "没有空白行或列需要清理！")
            return
        for r in reversed(empty_rows):
            self.table.removeRow(r)
        for c in reversed(empty_cols):
            self.table.removeColumn(c)
        new_formula = {}
        new_result = {}
        new_styles = {}
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                k = (r, c)
                if k in self.cell_formula:
                    new_formula[k] = self.cell_formula[k]
                if k in self.cell_result:
                    new_result[k] = self.cell_result[k]
                if k in self.cell_styles:
                    new_styles[k] = self.cell_styles[k]
        self.cell_formula = new_formula
        self.cell_result = new_result
        self.cell_styles = new_styles
        self.modified = True
        self.modified_label.setText("*")
        QMessageBox.information(self, "成功", f"清理完成！\n删除了 {len(empty_rows)} 行空白行和 {len(empty_cols)} 列空白列！")

    def get_symbol_config(self):
        if self.main_window and hasattr(self.main_window, 'get_symbol_config'):
            cfg = self.main_window.get_symbol_config()
            if cfg:
                return cfg
        return {}

    def symbol_add_func(self, sym, pos):
        if not self._check_pro_edit_permission():
            return
        items = self.table.selectedItems()
        if not items:
            QMessageBox.warning(self, "提示", "请先选择单元格")
            return
        cfg = self.get_symbol_config()
        if sym not in cfg:
            return
        left, right = cfg[sym][0], cfg[sym][1] if len(cfg[sym]) >= 2 else cfg[sym][0]
        self.table.blockSignals(True)
        try:
            for item in items:
                r, c = self.table.row(item), self.table.column(item)
                if self.cell_formula.get((r, c), ""):
                    continue
                old = self.cell_result.get((r, c), "")
                if self.skip_empty_cells and not old.strip():
                    continue
                if pos == "head":
                    new = left + old
                elif pos == "tail":
                    new = old + right
                elif pos == "both":
                    new = left + old + right
                elif pos == "split":
                    chars = list(old)
                    if self.split_mode == 0:
                        new = left.join(chars)
                    elif self.split_mode == 1:
                        new = "".join([f"{left}{c}{right}" for c in chars])
                    elif self.split_mode == 2:
                        new = ",".join([f"{left}{c}{right}" for c in chars])
                    else:
                        new = old
                else:
                    new = old
                self.push_undo_action(r, c, old, new)
                self.cell_result[(r, c)] = new
                item.setText(new)
                self.edited_cells.add((r, c))
                # 不再修改背景
                self.modified = True
                self.modified_label.setText("*")
        finally:
            self.table.blockSignals(False)
        QMessageBox.information(self, "成功", "处理完成！")

    def process_and_update_for_selected(self, op, pos):
        self.symbol_add_func(op, pos)

    def quick_process_with(self, sym, pos):
        self.now_sym = sym
        self.now_pos = pos
        self.process_and_update_for_selected(sym, pos)

    def fast_process_with_current_config(self):
        if not self._check_pro_edit_permission():
            return
        if not self.now_sym or not self.now_pos:
            QMessageBox.warning(self, "提示", "请先在符号处理菜单中选择符号和插入方式！")
            return
        self.process_and_update_for_selected(self.now_sym, self.now_pos)

    def show_context_menu(self, pos):
        if not self.is_pro:
            menu = QMenu()
            copy_action = menu.addAction("复制")
            copy_action.triggered.connect(self._copy_selected_cells)
            menu.exec_(self.table.viewport().mapToGlobal(pos))
            return
        menu = QMenu(self)
        if self.theme_manager:
            theme = self.theme_manager.get_theme()
            menu.setStyleSheet(f"QMenu {{ background-color: {theme['card']}; border: 1px solid {theme['border']}; border-radius: 8px; color: {theme['text']}; }}")
        if self.now_sym and self.now_pos:
            cfg = self.get_symbol_config()
            sym_info = cfg.get(self.now_sym, ("", "", "", ""))
            sym_label = sym_info[3] if len(sym_info) >= 4 else self.now_sym
            pos_map = {"head": "头部", "tail": "尾部", "both": "两端", "split": "分隔"}
            pos_label = pos_map.get(self.now_pos, self.now_pos)
            act = QAction(f"⚡ 快速应用: {sym_label} + {pos_label}", menu)
            act.setIcon(qta.icon('fa5s.bolt', color='#409eff'))
            act.setToolTip("使用当前选中的符号和位置快速应用到所选单元格")
            act.triggered.connect(self.fast_process_with_current_config)
            menu.addAction(act)
            menu.addSeparator()
        sym_menu = self._build_symbol_submenu()
        if sym_menu:
            menu.addMenu(sym_menu)
        a_sql = QAction(" 生成SQL IN语句", self)
        a_sql.setIcon(qta.icon('fa5s.database', color='#409eff'))
        a_sql.setToolTip("将所选单元格内容生成 SQL IN 语句")
        a_sql.triggered.connect(self.generate_sql_in_for_selected)
        menu.addAction(a_sql)
        a_copy = QAction(" 仅复制处理结果", self)
        a_copy.setIcon(qta.icon('fa5s.copy', color='#409eff'))
        a_copy.setToolTip("处理所选单元格并复制结果")
        a_copy.triggered.connect(self.process_and_copy_for_selected)
        menu.addAction(a_copy)
        menu.addSeparator()
        copy_act = QAction(" 复制", self)
        copy_act.setIcon(qta.icon('fa5s.copy', color='#606266'))
        copy_act.setShortcut(QKeySequence.Copy)
        copy_act.triggered.connect(self.copy_cell)
        menu.addAction(copy_act)
        if self.is_pro:
            paste_act = QAction(" 粘贴", self)
            paste_act.setIcon(qta.icon('fa5s.paste', color='#606266'))
            paste_act.setShortcut(QKeySequence.Paste)
            paste_act.triggered.connect(self.paste_cell)
            menu.addAction(paste_act)
            clear_act = QAction(" 清除内容", self)
            clear_act.setIcon(qta.icon('fa5s.eraser', color='#606266'))
            clear_act.triggered.connect(self.clear_selected)
            menu.addAction(clear_act)
            menu.addSeparator()
            add_row_act = QAction(" 在下方插入行", self)
            add_row_act.setIcon(qta.icon('fa5s.plus', color='#409eff'))
            add_row_act.triggered.connect(self.insert_row_below)
            menu.addAction(add_row_act)
            del_row_act = QAction(" 删除行", self)
            del_row_act.setIcon(qta.icon('fa5s.trash-alt', color='#dc2626'))
            del_row_act.triggered.connect(self.delete_row)
            menu.addAction(del_row_act)
            menu.addSeparator()
            clean_act = QAction(" 自动清理空白行/列", self)
            clean_act.setIcon(qta.icon('fa5s.broom', color='#409eff'))
            clean_act.triggered.connect(self.clean_empty_rows_columns)
            menu.addAction(clean_act)
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def generate_sql_in_for_selected(self):
        if not self.is_pro:
            return
        items = self.table.selectedItems()
        if not items:
            QMessageBox.warning(self, "提示", "请先选择单元格")
            return
        lines = [it.text().strip() for it in items if it.text().strip()]
        if not lines:
            QMessageBox.warning(self, "提示", "选中的单元格为空")
            return
        quoted = [f"'{l}'" for l in lines]
        sql_in = "IN (" + ", ".join(quoted) + ")"
        QApplication.clipboard().setText(sql_in)
        QMessageBox.information(self, "成功", f"SQL IN语句已复制到剪贴板！\n\n{sql_in}")

    def process_and_copy_for_selected(self):
        if not self.is_pro:
            return
        items = self.table.selectedItems()
        if not items:
            QMessageBox.warning(self, "提示", "请先选择单元格")
            return
        text_data = "\n".join([it.text() for it in items])
        self.process_and_copy(text_data)

    def process_and_copy(self, text):
        if not self.is_pro:
            return
        processed = self.process_text(text, "double_quote", "split")
        QApplication.clipboard().setText(processed)
        QMessageBox.information(self, "成功", "处理结果已复制到剪贴板！")

    def process_text(self, text, op, pos):
        if not self.is_pro:
            return text
        pure = text.strip()
        if not pure:
            return text
        cfg = self.get_symbol_config()
        if op not in cfg:
            return text
        left, right = cfg[op][0], cfg[op][1] if len(cfg[op]) >= 2 else cfg[op][0]
        if pos == "head":
            return left + pure
        elif pos == "tail":
            return pure + right
        elif pos == "both":
            return left + pure + right
        elif pos == "split":
            chars = list(pure)
            if self.split_mode == 0:
                return left.join(chars)
            elif self.split_mode == 1:
                return "".join([f"{left}{c}{right}" for c in chars])
            elif self.split_mode == 2:
                return ",".join([f"{left}{c}{right}" for c in chars])
        return pure

    def copy_cell(self):
        item = self.table.currentItem()
        if item:
            QApplication.clipboard().setText(item.text())

    def paste_cell(self):
        if not self._check_pro_edit_permission():
            return
        text = QApplication.clipboard().text()
        r, c = self.table.currentRow(), self.table.currentColumn()
        if r >= 0 and c >= 0:
            item = self.table.item(r, c) or QTableWidgetItem()
            self.table.setItem(r, c, item)
            item.setText(text)
            self.edited_cells.add((r, c))
            self.modified = True
            self.modified_label.setText("*")

    def clear_selected(self):
        if not self._check_pro_edit_permission():
            return
        for item in self.table.selectedItems():
            old = item.text()
            r, c = self.table.row(item), self.table.column(item)
            self.push_undo_action(r, c, old, "")
            item.setText("")
            self.edited_cells.add((r, c))
            self.modified = True
            self.modified_label.setText("*")

    def on_cell_double_clicked(self):
        if not self.is_pro:
            return
        cur = self.table.currentItem()
        if not cur:
            return
        r, c = cur.row(), cur.column()
        f = self.cell_formula.get((r, c), "")
        if f:
            cur.setText(f)

    def on_cell_changed(self, item):
        if not self.is_pro or self._in_undo_redo or self._is_loading_data or not item:
            return
        r, c = item.row(), item.column()
        txt = item.text().strip()
        old = self.cell_result.get((r, c), "")
        if old == "" and item.data(Qt.UserRole):
            old = str(item.data(Qt.UserRole))
        if old != txt:
            self.push_undo_action(r, c, old, txt)
            if txt.startswith("="):
                self.cell_formula[(r, c)] = txt
                calc_result = self._calculate_formula(txt)
                if calc_result != txt:
                    self.cell_result[(r, c)] = calc_result
                    item.setText(calc_result)
                else:
                    self.cell_result[(r, c)] = txt
                    item.setText(txt)
            else:
                self.cell_formula[(r, c)] = ""
                self.cell_result[(r, c)] = txt
                item.setText(txt)
            self.modified = True
            self.modified_label.setText("*")
            self.edited_cells.add((r, c))

    def confirm_formula_edit(self):
        if not self._check_pro_edit_permission():
            return
        item = self.table.currentItem()
        if not item:
            return
        r, c = item.row(), item.column()
        new = self.formula_edit.text().strip()
        try:
            if new.startswith("="):
                stack = []
                for ch in new:
                    if ch in '([':
                        stack.append(ch)
                    elif ch in ')]':
                        if not stack:
                            raise ValueError("括号不匹配")
                        stack.pop()
                if stack:
                    raise ValueError("括号不匹配")
                self.cell_formula[(r, c)] = new
                calc_result = self._calculate_formula(new)
                if calc_result != new:
                    self.cell_result[(r, c)] = calc_result
                    item.setText(calc_result)
                else:
                    self.cell_result[(r, c)] = new
                    item.setText(new)
            else:
                self.cell_formula[(r, c)] = ""
                self.cell_result[(r, c)] = new
                item.setText(new)
            self.modified = True
            self.modified_label.setText("*")
            self.edited_cells.add((r, c))
            self.formula_edit.clearFocus()
            self.table.setFocus()
        except Exception as e:
            QMessageBox.warning(self, "公式错误", f"公式处理失败：{str(e)}")
            oldv = self.cell_result.get((r, c), "")
            self.formula_edit.setText(oldv)
            item.setText(oldv)

    def _parse_cell_ref(self, ref):
        import re
        m = re.match(r'([A-Za-z]+)([0-9]+)', ref)
        if not m:
            return None, None
        alpha, num = m.group(1), m.group(2)
        col = 0
        for ch in alpha.upper():
            col = col * 26 + (ord(ch) - ord('A') + 1)
        col -= 1
        row = int(num) - 1
        return row, col

    def _get_cell_value(self, row, col):
        if 0 <= row < self.table.rowCount() and 0 <= col < self.table.columnCount():
            item = self.table.item(row, col)
            if item:
                val = item.text()
                try:
                    return float(val) if '.' in val else int(val)
                except:
                    return val
        return None

    def push_undo_action(self, row, col, old_val, new_val, extra=None):
        if self._in_undo_redo:
            return
        action = {'row': row, 'col': col, 'old_value': old_val, 'new_value': new_val}
        if extra:
            action['extra'] = extra
        self.undo_stack.append(action)
        if len(self.undo_stack) > self.MAX_UNDO_STEPS:
            self.undo_stack.pop(0)
        self.redo_stack.clear()

    def undo(self):
        if not self.undo_stack:
            return
        self._in_undo_redo = True
        act = self.undo_stack.pop()
        
        # 处理合并/拆分操作
        if 'extra' in act and act['extra'].get('type') in ['merge', 'split']:
            self._undo_merge_split(act, is_undo=True)
        else:
            r, c = act['row'], act['col']
            item = self.table.item(r, c)
            if item:
                self.redo_stack.append({'row': r, 'col': c, 'old_value': act['old_value'], 'new_value': item.text()})
                if act['old_value'] and act['old_value'].startswith('='):
                    self.cell_formula[(r, c)] = act['old_value']
                    res = self._calculate_formula(act['old_value'])
                    self.cell_result[(r, c)] = res if res != act['old_value'] else act['old_value']
                    item.setText(self.cell_result[(r, c)])
                else:
                    self.cell_formula[(r, c)] = ""
                    self.cell_result[(r, c)] = act['old_value'] or ""
                    item.setText(act['old_value'] or "")
        
        self._in_undo_redo = False
        self.modified = True
        self.modified_label.setText("*")

    def redo(self):
        if not self.redo_stack:
            return
        self._in_undo_redo = True
        act = self.redo_stack.pop()
        
        # 处理合并/拆分操作
        if 'extra' in act and act['extra'].get('type') in ['merge', 'split']:
            self._undo_merge_split(act, is_undo=False)
        else:
            r, c = act['row'], act['col']
            item = self.table.item(r, c)
            if item:
                self.undo_stack.append({'row': r, 'col': c, 'old_value': act['old_value'], 'new_value': item.text()})
                if act['new_value'] and act['new_value'].startswith('='):
                    self.cell_formula[(r, c)] = act['new_value']
                    res = self._calculate_formula(act['new_value'])
                    self.cell_result[(r, c)] = res if res != act['new_value'] else act['new_value']
                    item.setText(self.cell_result[(r, c)])
                else:
                    self.cell_formula[(r, c)] = ""
                    self.cell_result[(r, c)] = act['new_value'] or ""
                    item.setText(act['new_value'] or "")
        
        self._in_undo_redo = False
        self.modified = True
        self.modified_label.setText("*")
    
    def _undo_merge_split(self, action, is_undo=True):
        """撤销/重做合并或拆分操作"""
        extra = action['extra']
        action_type = extra['type']
        
        if is_undo:
            # 撤销：恢复到操作前状态
            if action_type == 'merge':
                # 撤销合并：恢复单元格内容和合并状态
                self.merged_cells = extra['old_merged']
                for (r, c), text in extra['cell_contents'].items():
                    self.table.setSpan(r, c, 1, 1)
                    item = self.table.item(r, c)
                    if item:
                        item.setText(text)
                # 重新应用之前的合并
                for (mr1, mc1, mr2, mc2) in self.merged_cells:
                    self.table.setSpan(mr1, mc1, mr2 - mr1 + 1, mc2 - mc1 + 1)
            else:  # split
                # 撤销拆分：重新合并单元格
                self.merged_cells = extra['old_merged']
                top, left, bottom, right = extra['top'], extra['left'], extra['bottom'], extra['right']
                for r in range(top, bottom+1):
                    for c in range(left, right+1):
                        self.table.setSpan(r, c, 1, 1)
                self.table.setSpan(top, left, bottom - top + 1, right - left + 1)
                # 清空其他单元格内容
                for r in range(top, bottom+1):
                    for c in range(left, right+1):
                        if r == top and c == left:
                            continue
                        item = self.table.item(r, c)
                        if item:
                            item.setText("")
                self.merged_cells.append((top, left, bottom, right))
            
            # 添加到redo栈
            self.redo_stack.append(action)
        else:
            # 重做：重新执行操作
            if action_type == 'merge':
                # 重新合并
                top, left, bottom, right = extra['top'], extra['left'], extra['bottom'], extra['right']
                for r in range(top, bottom+1):
                    for c in range(left, right+1):
                        self.table.setSpan(r, c, 1, 1)
                new_merged = []
                for (mr1, mc1, mr2, mc2) in self.merged_cells:
                    if not (max(mr1, top) <= min(mr2, bottom) and max(mc1, left) <= min(mc2, right)):
                        new_merged.append((mr1, mc1, mr2, mc2))
                self.merged_cells = new_merged
                self.table.setSpan(top, left, bottom - top + 1, right - left + 1)
                main_item = self.table.item(top, left)
                main_text = main_item.text() if main_item else ""
                for r in range(top, bottom+1):
                    for c in range(left, right+1):
                        if r == top and c == left:
                            continue
                        item = self.table.item(r, c)
                        if item:
                            item.setText("")
                self.merged_cells.append((top, left, bottom, right))
            else:  # split
                # 重新拆分
                top, left, bottom, right = extra['top'], extra['left'], extra['bottom'], extra['right']
                self.table.setSpan(top, left, 1, 1)
                new_merged = []
                for (mr1, mc1, mr2, mc2) in self.merged_cells:
                    if not (mr1 <= top <= mr2 and mc1 <= left <= mc2):
                        new_merged.append((mr1, mc1, mr2, mc2))
                self.merged_cells = new_merged
            
            # 添加到undo栈
            self.undo_stack.append(action)

    def create_shortcuts(self):
        from PySide6.QtGui import QShortcut, QKeySequence
        QShortcut(QKeySequence("Ctrl+Z"), self).activated.connect(self.undo)
        QShortcut(QKeySequence("Ctrl+Shift+Z"), self).activated.connect(self.redo)
        QShortcut(QKeySequence("Ctrl+S"), self).activated.connect(self.save_excel)
        QShortcut(QKeySequence("Ctrl+O"), self).activated.connect(self.import_excel)
        QShortcut(QKeySequence("Ctrl+F"), self).activated.connect(self.show_find_replace)

    def select_cell_change(self):
        item = self.table.currentItem()
        if not item:
            return
        r, c = item.row(), item.column()
        self.cell_address_label.setText(f"{chr(65+c)}{r+1}")
        f = self.cell_formula.get((r, c), "")
        if f:
            self.formula_edit.setText(f)
        else:
            self.formula_edit.setText(self.cell_result.get((r, c), ""))

    def update_cell_highlights(self):
        pass

    def set_language(self, lang):
        self.language = lang

    def set_dark_mode(self, is_dark):
        self.is_dark_mode = is_dark
        if self.theme_manager:
            self.apply_theme()
        else:
            self.apply_table_style_compat()

    def back_to_main(self):
        if self.modified and self.is_pro:
            reply = QMessageBox.question(self, "提示", "当前有未保存的修改，是否保存？",
                                         QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
            if reply == QMessageBox.Save:
                self.save_excel()
                if self.modified:
                    return
            elif reply == QMessageBox.Cancel:
                return
        if self.main_window:
            self.main_window.tab_widget.setCurrentIndex(0)
        self.close()

    def closeEvent(self, event):
        if self.modified and self.is_pro:
            reply = QMessageBox.question(self, "提示", "有未保存的修改，是否保存？",
                                         QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
            if reply == QMessageBox.Save:
                self.save_excel()
                event.accept()
            elif reply == QMessageBox.Discard:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    def resizeEvent(self, event):
        super().resizeEvent(event)

    def _load_excel_data(self, file_path, sheet_name=None):
        try:
            from openpyxl import load_workbook
            from PySide6.QtGui import QBrush, QColor
            from openpyxl.utils import get_column_letter

            # 文件大小检查 (限制为100MB)
            max_size_mb = 100
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)
            
            if file_size_mb > max_size_mb:
                reply = QMessageBox.question(self, "文件过大", 
                    f"文件大小为 {file_size_mb:.1f} MB，超过了 {max_size_mb} MB 的建议限制。\n"
                    "打开大文件可能会导致程序卡顿或内存占用过高。\n"
                    "是否继续打开？",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No:
                    return (False, "用户取消打开")

            self._is_loading_data = True
            self.table.setUpdatesEnabled(False)
            self.table.blockSignals(True)

            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.cell_formula.clear()
            self.cell_result.clear()
            self.merged_cells = []
            self.cell_styles.clear()
            self.edited_cells.clear()

            self.original_workbook = load_workbook(file_path, data_only=False)
            if sheet_name is None:
                ws_form = self.original_workbook.active
                sheet_name = ws_form.title
            else:
                ws_form = self.original_workbook[sheet_name]

            wb_val = load_workbook(file_path, read_only=True, data_only=True)
            ws_val = wb_val[sheet_name]

            self.sheet_names = self.original_workbook.sheetnames
            self.file_path = file_path
            self._add_to_recent_files(file_path)

            self.sheet_combo.blockSignals(True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(self.sheet_names)
            self.sheet_combo.setEnabled(True)
            current_idx = self.sheet_names.index(sheet_name) if sheet_name in self.sheet_names else 0
            self.sheet_combo.setCurrentIndex(current_idx)
            self.sheet_combo.blockSignals(False)

            max_r = ws_form.max_row
            max_c = ws_form.max_column
            if max_c == 0:
                max_c = 1

            MAX_ALLOWED_ROWS = 50000
            if max_r > MAX_ALLOWED_ROWS:
                max_r = MAX_ALLOWED_ROWS
                QMessageBox.information(self, "提示", f"文件行数过多，已限制为前{MAX_ALLOWED_ROWS}行")

            self.merged_cells.clear()
            for merged_range in ws_form.merged_cells.ranges:
                min_col = merged_range.min_col
                min_row = merged_range.min_row
                max_col = merged_range.max_col
                max_row = merged_range.max_row
                tr1 = min_row - 1
                tc1 = min_col - 1
                tr2 = max_row - 1
                tc2 = max_col - 1
                if tr1 <= tr2 and tc1 <= tc2 and tr2 < MAX_ALLOWED_ROWS:
                    self.merged_cells.append((tr1, tc1, tr2, tc2))

            self.table.setColumnCount(max_c)
            self.table.setRowCount(max_r)
            headers = [get_column_letter(i+1) for i in range(max_c)]
            self.table.setHorizontalHeaderLabels(headers)

            for r in range(max_r):
                for c in range(max_c):
                    item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    self.table.setItem(r, c, item)

            row_idx = 0
            batch_size = 500
            for row_data in ws_val.iter_rows(min_row=1, max_row=max_r, max_col=max_c, values_only=True):
                tr = row_idx
                for col_idx, value in enumerate(row_data):
                    if col_idx >= max_c:
                        break
                    v_text = ""
                    if value is not None:
                        v_str = str(value)
                        if " 00:00:00" in v_str:
                            v_text = v_str.replace(" 00:00:00", "")
                        elif "T00:00:00" in v_str:
                            v_text = v_str.replace("T00:00:00", "")
                        else:
                            v_text = v_str
                    item = self.table.item(tr, col_idx)
                    if item is None:
                        item = QTableWidgetItem(v_text)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.table.setItem(tr, col_idx, item)
                    else:
                        item.setText(v_text)
                    self.cell_result[(tr, col_idx)] = v_text

                row_idx += 1
                if row_idx % batch_size == 0:
                    QApplication.processEvents()

            for r in range(1, min(max_r, ws_form.max_row)+1):
                for c in range(1, min(max_c, ws_form.max_column)+1):
                    tr, tc = r-1, c-1
                    cell_form = ws_form.cell(r, c)

                    has_special = (cell_form.has_style or
                                cell_form.data_type == "f" or
                                (isinstance(cell_form.value, str) and cell_form.value.startswith("=")) or
                                cell_form.font or cell_form.fill or cell_form.alignment)
                    if not has_special:
                        continue

                    f_str = ""
                    if cell_form.data_type == "f":
                        f_str = cell_form.value
                    if not f_str and isinstance(cell_form.value, str) and cell_form.value.startswith("="):
                        f_str = cell_form.value

                    if f_str:
                        self.cell_formula[(tr, tc)] = f_str
                        item = self.table.item(tr, tc)
                        if item:
                            item.setText(f_str)
                            item.setData(Qt.UserRole, f_str)
                            # 公式单元格不设置背景

                    # 样式 - 字体
                    if cell_form.font and (cell_form.font.bold or cell_form.font.italic or cell_form.font.underline):
                        font = QFont()
                        if cell_form.font.name:
                            font.setFamily(cell_form.font.name)
                        size = getattr(cell_form.font, 'size', None)
                        if size is not None and isinstance(size, (int, float)) and size > 0:
                            font.setPointSize(size)
                        bold = cell_form.font.bold if cell_form.font.bold is not None else False
                        font.setBold(bold)
                        italic = cell_form.font.italic if cell_form.font.italic is not None else False
                        font.setItalic(italic)
                        underline = cell_form.font.underline if cell_form.font.underline is not None else False
                        font.setUnderline(underline)
                        item = self.table.item(tr, tc)
                        if item:
                            item.setFont(font)

                    # 字体颜色
                    if cell_form.font and cell_form.font.color and cell_form.font.color.rgb:
                        try:
                            rgb = str(cell_form.font.color.rgb)
                            if rgb.startswith("FF"):
                                rgb = rgb[2:]
                            if len(rgb) >= 6:
                                item = self.table.item(tr, tc)
                                if item:
                                    item.setForeground(QBrush(QColor(f"#{rgb}")))
                        except:
                            pass

                    # 背景填充（保留原始背景，不覆盖）
                    if cell_form.fill and cell_form.fill.fgColor and cell_form.fill.fgColor.rgb:
                        try:
                            rgb = str(cell_form.fill.fgColor.rgb)
                            if rgb.startswith("FF"):
                                rgb = rgb[2:]
                            if len(rgb) >= 6 and rgb != "000000":
                                item = self.table.item(tr, tc)
                                if item:
                                    item.setBackground(QBrush(QColor(f"#{rgb}")))
                        except:
                            pass

                    if cell_form.alignment:
                        item = self.table.item(tr, tc)
                        if item:
                            if cell_form.alignment.horizontal in ('center', 'centerContinuous'):
                                item.setTextAlignment(Qt.AlignCenter)
                            elif cell_form.alignment.horizontal == 'right':
                                item.setTextAlignment(Qt.AlignRight)
                            else:
                                item.setTextAlignment(Qt.AlignLeft)

                    self.cell_styles[(tr, tc)] = {
                        'font': cell_form.font,
                        'fill': cell_form.fill,
                        'border': cell_form.border,
                        'alignment': cell_form.alignment,
                        'number_format': cell_form.number_format
                    }

                if r % 50 == 0:
                    QApplication.processEvents()
                    if self._is_loading_data == False:
                        return False

            for idx, (tr1, tc1, tr2, tc2) in enumerate(self.merged_cells):
                if 0 <= tr1 <= tr2 < self.table.rowCount() and 0 <= tc1 <= tc2 < self.table.columnCount():
                    row_span = tr2 - tr1 + 1
                    col_span = tc2 - tc1 + 1
                    self.table.setSpan(tr1, tc1, row_span, col_span)
                    main_item = self.table.item(tr1, tc1)
                    if main_item:
                        main_item.setToolTip(main_item.text())
                if idx % 100 == 0:
                    QApplication.processEvents()

            for c in range(1, max_c+1):
                col_letter = get_column_letter(c)
                if col_letter in ws_form.column_dimensions and ws_form.column_dimensions[col_letter].width:
                    width = ws_form.column_dimensions[col_letter].width
                    if width > 0:
                        self.table.setColumnWidth(c-1, int(width * 7))
                if c % 50 == 0:
                    QApplication.processEvents()
            for r in range(1, max_r+1):
                if r in ws_form.row_dimensions and ws_form.row_dimensions[r].height:
                    height = ws_form.row_dimensions[r].height
                    if height > 0:
                        self.table.setRowHeight(r-1, int(height * 4 / 3))
                if r % 50 == 0:
                    QApplication.processEvents()

            wb_val.close()
            # 注意：self.original_workbook 是普通模式加载的工作簿，没有 close() 方法
            # Python 的垃圾回收会自动处理文件句柄，无需手动关闭

            self.table.verticalHeader().setDefaultSectionSize(24)
            self.table.horizontalHeader().setDefaultSectionSize(90)
            self.table.setShowGrid(True)
            self.table.setSelectionBehavior(QAbstractItemView.SelectItems)

            self.btn_save.setEnabled(self.is_pro)
            self.btn_add.setEnabled(self.is_pro)
            self.btn_delete.setEnabled(self.is_pro)
            self.btn_clean.setEnabled(self.is_pro)
            self.btn_export_csv.setEnabled(self.is_pro)
            self.btn_find.setEnabled(self.is_pro)
            self.btn_freeze_row.setEnabled(self.is_pro)
            self.btn_zoom_in.setEnabled(True)
            self.btn_zoom_out.setEnabled(True)
            self.btn_zoom_reset.setEnabled(True)

            self.modified = False
            self.modified_label.setText("")
            self.sheet_label.setText(f"Sheet: {sheet_name}")

            self._save_original_fonts()
            
            # 加载完数据后，设置基准值
            if self.table.rowCount() > 0 and self.table.columnCount() > 0:
                item = self.table.item(0, 0)
                if item:
                    self._base_font_size = item.font().pointSize() or 10
                else:
                    self._base_font_size = 10
            else:
                self._base_font_size = 10
            
            # 基准行高列宽
            self._base_row_height = 25
            self._base_col_width = 100
            
            fd, self.backup_file = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            return True, None

        except Exception as e:
            import traceback
            error_msg = f"加载Excel文件失败: {str(e)}"
            traceback.print_exc()
            try:
                if 'wb_val' in locals() and wb_val:
                    wb_val.close()
                # 注意：self.original_workbook 是普通模式加载的工作簿，没有 close() 方法
                # Python 的垃圾回收会自动处理文件句柄，无需手动关闭
            except:
                pass
            QMessageBox.critical(self, "加载失败", error_msg)
            return False, error_msg
        finally:
            self._is_loading_data = False
            self.table.setUpdatesEnabled(True)
            self.table.blockSignals(False)

    def _calculate_formula(self, formula):
        if not formula.startswith("="):
            return formula
        expr = formula[1:].strip()
        try:
            import re
            
            # 先处理所有函数，直接在参数内部计算，而不预先替换单元格
            def evaluate_function(match):
                func_name = match.group(1).upper()
                args_str = match.group(2)
                
                if func_name in ['SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN']:
                    # 直接评估范围或列表，不进行单元格预替换
                    values = self._eval_range_or_list(args_str)
                    if func_name == 'SUM':
                        return str(sum(values)) if values else "0"
                    elif func_name == 'AVERAGE':
                        return str(sum(values) / len(values)) if values else "0"
                    elif func_name == 'COUNT':
                        return str(len([v for v in values if isinstance(v, (int, float))]))
                    elif func_name == 'MAX':
                        return str(max(values)) if values else "0"
                    elif func_name == 'MIN':
                        return str(min(values)) if values else "0"
                elif func_name == 'IF':
                    parts = self._parse_if_args(args_str)
                    if len(parts) >= 2:
                        cond, true_val = parts[0], parts[1]
                        false_val = parts[2] if len(parts) > 2 else ""
                        cond_result = self._eval_condition(cond.strip())
                        if cond_result:
                            return self._eval_atom_str(true_val.strip())
                        else:
                            return self._eval_atom_str(false_val.strip())
                    return "#ERROR"
                elif func_name == 'VLOOKUP':
                    parts = args_str.split(',')
                    if len(parts) >= 3:
                        lookup_value = parts[0].strip()
                        table_array = parts[1].strip()
                        try:
                            col_index = int(parts[2].strip()) - 1
                            table_data = self._eval_range(table_array)
                            lookup_val = self._eval_atom(lookup_value)
                            for row in table_data:
                                if row and len(row) > 0 and row[0] == lookup_val:
                                    if col_index < len(row):
                                        return str(row[col_index])
                        except:
                            pass
                    return "#N/A"
                return match.group(0)
            
            # 从内向外替换函数
            def replace_functions(s):
                result = s
                func_pattern = r'([A-Za-z]+)\(([^)]+)\)'
                while True:
                    new_result = re.sub(func_pattern, evaluate_function, result, flags=re.IGNORECASE)
                    if new_result == result:
                        break
                    result = new_result
                return result
            
            # 先处理所有函数，将它们替换为计算结果
            expr = replace_functions(expr)
            
            # 然后只替换剩余的单个单元格引用（非函数参数，因为函数已经处理完了）
            def repl_cell(m):
                ref = m.group(0)
                val = self._get_cell_value_by_ref(ref)
                return str(val) if val is not None else "0"
            
            expr = re.sub(r'\b([A-Za-z]+[0-9]+)\b', repl_cell, expr)
            
            result = self._safe_calculate(expr)
            if isinstance(result, (int, float)):
                if isinstance(result, float) and result.is_integer():
                    result = int(result)
                return str(result)
            return str(result)
        except Exception:
            return formula
    
    def _parse_if_args(self, args_str):
        args = []
        current = ""
        depth = 0
        for char in args_str:
            if char == '(' and depth == 0:
                current += char
                depth += 1
            elif char == ')':
                depth -= 1
                if depth < 0:
                    break
                current += char
            elif char == ',' and depth == 0:
                args.append(current)
                current = ""
            else:
                current += char
        if current:
            args.append(current)
        return args
    
    def _eval_atom_str(self, atom):
        result = self._eval_atom(atom)
        if result is None:
            return ""
        return str(result)

    def _safe_calculate(self, expr):
        import re
        if not re.fullmatch(r'[0-9+\-*/().\s]+', expr):
            raise ValueError("非法表达式")
        forbidden = ['__', 'import', 'eval', 'exec', 'compile', 'open', 'file', 'globals', 'locals', 'getattr', 'setattr', 'delattr', 'hasattr']
        for f in forbidden:
            if f in expr.lower():
                raise ValueError("非法表达式")
        def tokenize(s):
            tokens = []
            i = 0
            n = len(s)
            while i < n:
                if s[i].isspace():
                    i += 1
                elif s[i] in '+-*/()':
                    tokens.append(s[i])
                    i += 1
                elif s[i].isdigit() or s[i] == '.':
                    j = i
                    while j < n and (s[j].isdigit() or s[j] == '.'):
                        j += 1
                    tokens.append(s[i:j])
                    i = j
                else:
                    raise ValueError("非法字符")
            return tokens
        def eval_atom(tokens, idx):
            if idx >= len(tokens):
                raise ValueError("意外结束")
            token = tokens[idx]
            if token == '(':
                idx += 1
                val, idx = eval_expr(tokens, idx)
                if idx >= len(tokens) or tokens[idx] != ')':
                    raise ValueError("缺少右括号")
                idx += 1
                return val, idx
            elif token in '+-':
                sign = 1 if token == '+' else -1
                idx += 1
                val, idx = eval_atom(tokens, idx)
                return sign * val, idx
            else:
                try:
                    val = float(token)
                    idx += 1
                    return val, idx
                except:
                    raise ValueError("非法数字")
        def eval_term(tokens, idx):
            val, idx = eval_atom(tokens, idx)
            while idx < len(tokens) and tokens[idx] in '*/':
                op = tokens[idx]
                idx += 1
                right, idx = eval_atom(tokens, idx)
                if op == '*':
                    val *= right
                else:
                    if right == 0:
                        raise ValueError("除零错误")
                    val /= right
            return val, idx
        def eval_expr(tokens, idx):
            val, idx = eval_term(tokens, idx)
            while idx < len(tokens) and tokens[idx] in '+-':
                op = tokens[idx]
                idx += 1
                right, idx = eval_term(tokens, idx)
                if op == '+':
                    val += right
                else:
                    val -= right
            return val, idx
        tokens = tokenize(expr)
        if not tokens:
            return 0
        result, idx = eval_expr(tokens, 0)
        if idx != len(tokens):
            raise ValueError("表达式未完全解析")
        return result

    def _eval_range_or_list(self, range_str):
        if ':' in range_str:
            return self._eval_range(range_str)
        else:
            items = [s.strip() for s in range_str.split(',')]
            values = []
            for item in items:
                val = self._eval_atom(item)
                if val is not None:
                    values.append(val)
            return values

    def _eval_range(self, range_str):
        start_ref, end_ref = range_str.split(':')
        start_row, start_col = self._parse_cell_ref(start_ref)
        end_row, end_col = self._parse_cell_ref(end_ref)
        if start_row is None or end_row is None:
            return []
        values = []
        for r in range(min(start_row, end_row), max(start_row, end_row) + 1):
            row_vals = []
            for c in range(min(start_col, end_col), max(start_col, end_col) + 1):
                val = self._get_cell_value(r, c)
                row_vals.append(val)
            values.append(row_vals)
        flat = []
        for row in values:
            flat.extend(row)
        return flat

    def _eval_atom(self, atom):
        atom = atom.strip()
        if atom.startswith('"') and atom.endswith('"'):
            return atom[1:-1]
        try:
            if '.' in atom:
                return float(atom)
            else:
                return int(atom)
        except ValueError:
            val = self._get_cell_value_by_ref(atom)
            return val if val is not None else atom

    def _eval_condition(self, cond):
        import re
        ops = ['>=', '<=', '>', '<', '==', '!=']
        for op in ops:
            if op in cond:
                left, right = cond.split(op, 1)
                left_val = self._eval_atom(left.strip())
                right_val = self._eval_atom(right.strip())
                if op == '>=':
                    return left_val >= right_val
                elif op == '<=':
                    return left_val <= right_val
                elif op == '>':
                    return left_val > right_val
                elif op == '<':
                    return left_val < right_val
                elif op == '==':
                    return left_val == right_val
                elif op == '!=':
                    return left_val != right_val
        return False

    def _get_cell_value_by_ref(self, ref):
        row, col = self._parse_cell_ref(ref)
        if row is not None:
            return self._get_cell_value(row, col)
        return None

    def auto_backup(self):
        if (self.modified and self.file_path and self.is_pro and 
            self.backup_file and isinstance(self.backup_file, str) and 
            self.original_workbook is not None):
            try:
                self.save_excel(self.backup_file)
            except Exception as e:
                print(f"自动备份失败: {e}")

    def _clear_all_highlights(self):
        # 不清除任何背景，仅清除记录
        self.edited_cells.clear()

    def new_file(self):
        if not self.is_pro:
            QMessageBox.warning(self, "权限不足", "新建文件为 PRO 版专属")
            return
        if self.modified:
            reply = QMessageBox.question(self, "提示", "当前内容未保存，是否保存？",
                                         QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
            if reply == QMessageBox.Save:
                self.save_excel()
                if self.modified:
                    return
            elif reply == QMessageBox.Cancel:
                return
        self._reset_to_initial_state()

    def load_sheet_with_formula(self, ws):
        pass